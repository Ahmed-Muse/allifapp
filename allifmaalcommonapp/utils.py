# C:\am\allifapp\allifapperp\allifmaalcommonapp\utils.py

# C:\am\allifapp\allifapperp\allifmaalcommonapp\utils.py
from django import forms # Import forms for type hinting if needed
from django.db.models import QuerySet # Import QuerySet for type hinting
from django.utils import timezone 
from django.db.models import QuerySet, Model # Import Model
from django.http import HttpRequest
import datetime
from typing import Optional # Import Optional for type hinting
from .models import *
from allifmaalshaafiapp.models import *
from django.db.models import QuerySet, Model, Q # Ensure Q is imported for complex lookups
from django.db import IntegrityError

import logging

logger = logging.getLogger(__name__)
# --- Standard Python Imports ---
from typing import Optional, Dict, List, Any
import datetime 
from decimal import Decimal 
import io # For in-memory file handling
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill # For styling Excel
# --- Django Imports ---
from django.db.models import QuerySet, Model, Q 
from django.http import HttpRequest, Http404, HttpResponse 
from django.shortcuts import render, get_object_or_404, redirect 
from django.urls import reverse 
from django.utils import timezone 
from django.template.loader import get_template 
from django.db import transaction 
from django.contrib import messages 

# --- Third-party Imports ---
from xhtml2pdf import pisa 
import openpyxl # Import openpyxl for Excel generation



from typing import Optional, Dict, List, Any

# For PDF generation
from django.template.loader import get_template
from xhtml2pdf import pisa # Assuming xhtml2pdf is installed (pip install xhtml2pdf)
from django.shortcuts import render,redirect,get_object_or_404
from .allifutils import common_shared_data
from django.urls import reverse,NoReverseMatch
from django.http import Http404

from .models import CommonPurchaseOrdersModel, CommonPurchaseOrderItemsModel

# Define sort mappings and default sort fields for different models here.
# Use a clear identifier (e.g., model name in lowercase) as the key.
allif_model_sort_configs = {
    'commonexpensesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Quantity Ascending": "quantity",
            "Quantity Descending": "-quantity",
            "Amount Ascending": "amount",
            "Amount Descending": "-amount",
            "Created At Ascending": "date",
            "Created At Descending": "-date",
            "Status Ascending": "status",
            "Status Descending": "-status",
            "Supplier Name Ascending": "supplier__name", 
            "Supplier Name Descending": "-supplier__name",
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
    'commoncurrenciesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
    'commoncategoriesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
     'commoncodesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonbanksmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonshareholderbankdepositsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonbankwithdrawalsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
    
    'commonsuppliersmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commoncustomersmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     
      'commonassetsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      
      
      'commonexpensesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      
     'commontransactionsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonspaceunitsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonstocksmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
      'commonpurchaseordersmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      
       'commonstocktransferordersmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
      'commonquotesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
      'commoninvoicesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commoncreditnotesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      'commonledgerentriesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
    'commonsupplierpaymentsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
      'commoncustomerpaymentsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      'commonsalariesmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     'commonjobsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
      'commontransitmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
     'commonlogsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      
       'commondivisionsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
      
     
     
     
     
    'commoncompanyscopemodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
    
     'commonapproversmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
    'assessmentsmodel': { # Key should be consistent, e.g., model._meta.model_name
        'sort_mapping': {
            "Name Ascending": "name",
            "Name Descending": "-name",
            "Description Ascending": "description",
            "Description Descending": "-description",
            
           
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending', # A default label for initial load
    },
     
     
    
    
    'commontasksmodel': {
        'sort_mapping': {
            "Task Name Ascending": "name",
            "Task Name Descending": "-name",
            "Date Ascending": "date",
            "Date Descending": "-date",
            "Priority Ascending": "priority",
            "Priority Descending": "-priority",
            "Status Ascending": "status",
            "Status Descending": "-status",
            "Created At Ascending": "date",
            "Created At Descending": "-date",
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Date Descending',
    },
    'triagesmodel': { # Assuming TriagesModel exists and is named this way
        'sort_mapping': {
            "Medical File Ascending": "medical_file",
            "Medical File Descending": "-medical_file",
            "Created At Ascending": "date",
            "Created At Descending": "-date",
            "Pain Level Ascending": "pain_level",
            "Pain Level Descending": "-pain_level",
            "Weight Ascending": "weight",
            "Weight Descending": "-weight",
            "Height Ascending": "height",
            "Height Descending": "-height",
            
             "Temperature Ascending": "temperature",
            "Temperature Descending": "-temperature",
        },
        'default_sort_field': '-date',
        'default_ui_label': 'Created At Descending',
    }
    # Add other model sort configurations here
}




# --- NEW: Search Configuration Map ---
# Define which fields are searchable for each model
allif_search_config_mapping = {
    'CommonLogsModel': ['name__icontains', 'description__icontains','balance__icontains'],
    'CommonStocksModel': ['description__icontains', 'number__icontains','name__icontains','unitPrice__icontains'],
    'CommonCurrenciesModel': ['name__icontains', 'description__icontains'],
    'CommonCategoriesModel': ['name__icontains', 'description__icontains'],
    'CommonCodesModel': ['name__icontains', 'description__icontains'],
    'CommonChartofAccountsModel': ['name__icontains', 'description__icontains'],
    'CommonBanksModel': ['name__icontains', 'description__icontains','balance__icontains'],
    
    'CommonShareholderBankDepositsModel': ['name__icontains', 'description__icontains','amount__icontains'],
    
    'CommonBankWithdrawalsModel': ['name__icontains', 'description__icontains','amount__icontains'],
    'CommonSuppliersModel': ['name__icontains', 'description__icontains','balance__icontains'],
    
     'CommonCustomersModel': ['name__icontains', 'description__icontains','balance__icontains'],
    
     'CommonAssetsModel': ['name__icontains', 'description__icontains','balance__icontains'],
     'CommonTransactionsModel': ['name__icontains', 'description__icontains','balance__icontains'],
    
    'CommonSpaceUnitsModel': ['name__icontains', 'description__icontains','balance__icontains'],
  
    'CommonPurchaseOrdersModel': ['name__icontains', 'description__icontains','number__icontains'],
    
    'CommonQuotesModel': ['name__icontains', 'description__icontains','number__icontains'],
    'CommonStockTransferOrdersModel': ['name__icontains', 'description__icontains','number__icontains'],
    
    'CommonInvoicesModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonCreditNotesModel': ['name__icontains', 'description__icontains','number__icontains'],
     
     'CommonLedgerEntriesModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonSupplierPaymentsModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonCustomerPaymentsModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonSalariesModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonJobsModel': ['name__icontains', 'description__icontains','number__icontains'],
     'CommonTransitModel': ['name__icontains', 'description__icontains','number__icontains'],
      'CommonProgressModel': ['name__icontains', 'description__icontains','number__icontains'],
     'TriagesModel': ['name__icontains', 'description__icontains','number__icontains'],
     'AssessmentsModel': ['name__icontains', 'description__icontains','number__icontains'],
     
    
    'CommonExpensesModel': ['description__icontains', 'amount__icontains', 'supplier__name__icontains'], # Example
    'CommonTasksModel': ['name__icontains', 'description__icontains'], # Example
    #'TriagesModel': ['medical_file__customer__icontains', 'pain_level__icontains'], # Example
    # Add configurations for other models as needed
}


# --- NEW: Advanced Search Configuration Map ---
# Define which fields are used for date and value range filtering for each model
allif_advanced_search_configs = {
    'CommonCurrenciesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Currency Name'},
        {'field': 'number', 'label': 'Currency Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Currency Description'},
        {'field': 'balance', 'label': 'Currency Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
    
    'CommonChartofAccountsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Currency Name'},
        {'field': 'number', 'label': 'Currency Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Currency Description'},
        {'field': 'balance', 'label': 'Currency Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
    
     'CommonShareholderBankDepositsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
      'CommonBankWithdrawalsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
      
    'CommonSuppliersModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
    
     'CommonCustomersModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     'CommonAssetsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
      'CommonTransactionsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
      
    'CommonStocksModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
    
    'CommonPurchaseOrdersModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
    
     'CommonStockTransferOrdersModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
      'CommonQuotesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
       'CommonInvoicesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
       'CommonCreditNotesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     'CommonLedgerEntriesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
     'CommonSupplierPaymentsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
     'CommonCustomerPaymentsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
      'CommonSalariesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
      
      'CommonJobsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
      
     'CommonTransitModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
     'CommonProgressModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
      'TriagesModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
       'AssessmentsModel': {
        'date_field': 'starts', # Name of the date field in CommonStocksModel
        'value_field': 'balance', # Name of the quantity/value field in CommonStocksModel
        'default_order_by_date': '-starts', # Default ordering for finding first/last date
        'default_order_by_value': '-balance', # Default ordering for finding largest value
        'excel_fields': [
        {'field': 'name', 'label': 'Name'},
        {'field': 'number', 'label': 'Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Description'},
        {'field': 'balance', 'label': 'Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
    
    
    },
     
     
     
     
     
     
     
     
    
    
    
    # Add configurations for other models here, e.g.:
    'CommonExpensesModel': {
        'date_field': 'starts',
         'value_field': 'balance',
         'default_order_by_date': '-starts',
         'default_order_by_value': '-balance',
         
          'excel_fields': [
            {'field': 'name', 'label': 'Currency Name'},
        {'field': 'number', 'label': 'Currency Number'},
        {'field': 'code', 'label': 'Code'},
        {'field': 'description', 'label': 'Currency Description'},
        {'field': 'balance', 'label': 'Currency Balance'},
        {'field': 'date', 'label': 'Date'},
        ]
     },
    
}




# --- ALLIF_MODEL_REGISTRY (Central Model Class Map) ---
# This dictionary maps model names (strings) to their actual Python model classes.
# It's used by all generic handlers (delete, search, advanced search, PDF generation)
# to dynamically get the correct model class based on a string name.
allif_main_models_registry = { 
    'CommonSectorsModel': CommonSectorsModel,
    'CommonCompanyScopeModel': CommonCompanyScopeModel,
    'CommonTaxParametersModel': CommonTaxParametersModel,
    'CommonExpensesModel': CommonExpensesModel,
    'CommonTasksModel': CommonTasksModel,
    'CommonBanksModel': CommonBanksModel,
    'TriagesModel': TriagesModel, 
    'CommonCurrenciesModel': CommonCurrenciesModel, # Corrected typo CommonCurrenciesModel
    'CommonSupplierPaymentsModel': CommonSupplierPaymentsModel,
    'CommonCustomerPaymentsModel': CommonCustomerPaymentsModel,
    'CommonPaymentTermsModel': CommonPaymentTermsModel,
    'CommonTransitModel': CommonTransitModel,
    'CommonStocksModel': CommonStocksModel, 
    'CommonPurchaseOrdersModel': CommonPurchaseOrdersModel, 
    'CommonPurchaseOrderItemsModel': CommonPurchaseOrderItemsModel, 
    'CommonExpensesModel': CommonExpensesModel,
    'CommonTasksModel': CommonTasksModel,
    'CommonBanksModel': CommonBanksModel,
    'CommonUnitsModel':CommonUnitsModel,
    'CommonCategoriesModel':CommonCategoriesModel,
    'CommonCodesModel':CommonCodesModel,
    'CommonApproversModel':CommonApproversModel,
    'CommonSupplierTaxParametersModel':CommonSupplierTaxParametersModel,
    'CommonChartofAccountsModel':CommonChartofAccountsModel,
    'CommonGeneralLedgersModel':CommonGeneralLedgersModel,
    'CommonShareholderBankDepositsModel':CommonShareholderBankDepositsModel,
    
    'CommonBankWithdrawalsModel':CommonBankWithdrawalsModel,
    'CommonSuppliersModel':CommonSuppliersModel,
    "CommonCustomersModel":CommonCustomersModel,
    "CommonAssetsModel":CommonAssetsModel,
    "CommonTransactionsModel":CommonTransactionsModel,
    "CommonTransactionItemsModel":CommonTransactionItemsModel,
    "CommonSpacesModel":CommonSpacesModel,
    "CommonSpaceUnitsModel":CommonSpaceUnitsModel,
    "CommonSpaceBookingItemsModel":CommonSpaceBookingItemsModel,
    "CommonPurchaseOrderMiscCostsModel":CommonPurchaseOrderMiscCostsModel,
    "CommonStockTransferOrderItemsModel":CommonStockTransferOrderItemsModel,
    "CommonStockTransferOrdersModel":CommonStockTransferOrdersModel,
    "CommonQuotesModel":CommonQuotesModel,
    "CommonQuoteItemsModel":CommonQuoteItemsModel,
    "CommonInvoicesModel":CommonInvoicesModel,
    "CommonInvoiceItemsModel":CommonInvoiceItemsModel,
    "CommonCreditNotesModel":CommonCreditNotesModel,
    "CommonCreditNoteItemsModel":CommonCreditNoteItemsModel,
    "CommonLedgerEntriesModel":CommonLedgerEntriesModel,
    "CommonSalariesModel":CommonSalariesModel,
    "CommonJobsModel":CommonJobsModel,
    "CommonJobItemsModel":CommonJobItemsModel,
    "CommonTransitModel":CommonTransitModel,
    "CommonTransitItemsModel":CommonTransitItemsModel,
   "CommonProgressModel":CommonProgressModel,
   "AssessmentsModel":AssessmentsModel,
   "CommonLogsModel":CommonLogsModel,
   "CommonDivisionsModel":CommonDivisionsModel,
  
   
}

# --- Document PDF Configuration Map ---
# Defines the main model, and optionally related items model and context keys for each document type
allif_main_document_pdf_configuration= {
    'CommonPurchaseOrdersModel': { # Key for Purchase Orders
        'main_model': 'CommonPurchaseOrdersModel', 
        'items_model': 'CommonPurchaseOrderItemsModel', # This is an optional key now
        'items_related_field': 'po_item_con', # This is an optional key now
        'title': 'Purchase Order',
        'filename_prefix': 'PO',
        'template_path': 'allifmaalcommonapp/ui/pdf/items-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        'related_lookups': ['supplier'], 
    },
    
     'CommonStockTransferOrdersModel': { # Key for Purchase Orders
        'main_model': 'CommonStockTransferOrdersModel', 
        'items_model': 'CommonStockTransferOrderItemsModel', # This is an optional key now
        'items_related_field': 'trans_ord_items_con', # This is an optional key now
        'title': 'Transfer Order',
        'filename_prefix': 'TRF',
        'template_path': 'allifmaalcommonapp/ui/pdf/items-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
     
     
      'CommonQuotesModel': { # Key for Purchase Orders
        'main_model': 'CommonQuotesModel', 
        'items_model': 'CommonQuoteItemsModel', # This is an optional key now
        'items_related_field': 'allifquoteitemconnector', # This is an optional key now
        'title': 'Transfer Order',
        'filename_prefix': 'TRF',
        'template_path': 'allifmaalcommonapp/quotes/quote-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
    'CommonInvoicesModel': { # Key for Purchase Orders
        'main_model': 'CommonInvoicesModel', 
        'items_model': 'CommonInvoiceItemsModel', # This is an optional key now
        'items_related_field': 'allifinvitemconnector', # This is an optional key now
        'title': 'Invoice',
        'filename_prefix': 'Inv',
        'template_path': 'allifmaalcommonapp/invoices/invoice-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
    
    'CommonSuppliersModel': { # Key for Purchase Orders
        'main_model': 'CommonSuppliersModel', 
        'items_model': 'CommonLedgerEntriesModel', # This is an optional key now
        'items_related_field': 'supplier', # This is an optional key now
        'title': 'Invoice',
        'filename_prefix': 'Inv',
        'template_path': 'allifmaalcommonapp/statements/suppliers/supplier_statement_pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
    
    'CommonCustomersModel': { # Key for Purchase Orders
        'main_model': 'CommonCustomersModel', 
        'items_model': 'CommonLedgerEntriesModel', # This is an optional key now
        'items_related_field': 'customer', # This is an optional key now
        'title': 'Invoice',
        'filename_prefix': 'Inv',
        'template_path': 'allifmaalcommonapp/statements/customers/customer_statement_pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
    
    
    
    
     
    'CommonCreditNotesModel': { # Key for Purchase Orders
        'main_model': 'CommonCreditNotesModel', 
        'items_model': 'CommonCreditNoteItemsModel', # This is an optional key now
        'items_related_field': 'credit_note', # This is an optional key now
        'title': 'Invoice',
        'filename_prefix': 'Inv',
        'template_path': 'allifmaalcommonapp/ui/pdf/items-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
   
    
    
    
    'CommonTransactionsModel': { # Key for Purchase Orders
        'main_model': 'CommonTransactionsModel', 
        'items_model': 'CommonSpaceBookingItemsModel', # This is an optional key now
        'items_related_field': 'trans_number', # This is an optional key now
        'title': 'Purchase Order',
        'filename_prefix': 'PO',
        'template_path': 'allifmaalcommonapp/booking/space_allocation_pdf.html', 
        'extra_context_map': { 
            'supplier': 'trans_number', 
        },
        'related_lookups': ['trans_number'], 
    },
    
    # Example for a model that has NO related items, just the main model's PDF 
    'CommonCurrenciesModel': {
        'main_model': 'CommonCurrenciesModel',
        # No 'items_model' or 'items_related_field' needed for this config
        'title': 'Currency Details',
        'filename_prefix': 'CUR',
        'template_path': 'allifmaalcommonapp/ui/pdf/item-pdf.html', # Create this template
        'extra_context_map': {}, # No extra fields to map
        'related_lookups': [], # No related lookups needed
    },
    
    'CommonTransactionsModel': {
        'main_model': 'CommonCurrenciesModel',
        # No 'items_model' or 'items_related_field' needed for this config
        'title': 'Currency Details',
        'filename_prefix': 'CUR',
        'template_path': 'allifmaalcommonapp/ui/pdf/item-pdf.html', # Create this template
        'extra_context_map': {}, # No extra fields to map
        'related_lookups': [], # No related lookups needed
    },
    
    
    'CommonJobsModel': { # Key for Purchase Orders
        'main_model': 'CommonJobsModel', 
        'items_model': 'CommonJobItemsModel', # This is an optional key now
        'items_related_field': 'credit_note', # This is an optional key now
        'title': 'Invoice',
        'filename_prefix': 'Inv',
        'template_path': 'allifmaalcommonapp/ui/pdf/items-pdf.html', 
        'extra_context_map': { 
            'supplier': 'po_supplier', 
        },
        #'related_lookups': ['supplier'], 
    },
    
     'CommonTransitModel': { # Key for Purchase Orders
        'main_model': 'CommonTransitModel', 
        'items_model': 'CommonTransitItemsModel', # This is an optional key now
        'items_related_field': 'shipment', # This is an optional key now
        'title': 'Transport',
        'filename_prefix': 'Shipment',
        'template_path': 'allifmaalcommonapp/transport/shipment_pdf.html', 
       
    },
    
    
    
}

# --- NEW: EXCEL_UPLOAD_CONFIGS (Centralized Excel Upload Configuration Map) ---
allif_excel_upload_configs= {
    'CommonCurrenciesModel': {
        'model': 'CommonCurrenciesModel', # String name for ALLIF_MODEL_REGISTRY lookup
        'required_excel_headers': ['Name', 'Code', 'Description'], # Headers that MUST be in Excel
        'field_mapping': { # Map Excel headers to model field names if different
            'Currency Name': 'name',
            'Code': 'code',
           
            'Description': 'description',
            #'Start Date': 'starts', # Assuming 'starts' is a DateTimeField/DateField
            #'End Date': 'ends',     # Assuming 'ends' is a DateTimeField/DateField
         
            #'Status': 'status', # Example status field
        },
        #'related_field_lookups': {
            # No related fields for CommonCurrencyModel in this example
        #},
        #'default_values': {
            #'is_active': True, # Default value if 'Is Active' column is missing or empty
            #'status': 'active', # Default status if 'Status' column is missing or empty
        #},
    },
    
     'CommonAssetsModel': {
        'model': 'CommonAssetsModel', # String name for ALLIF_MODEL_REGISTRY lookup
        'required_excel_headers': ['Name', 'Code', 'Description'], # Headers that MUST be in Excel
        'field_mapping': { # Map Excel headers to model field names if different
            'Currency Name': 'name',
            'Code': 'code',
           
            'Description': 'description',
            #'Start Date': 'starts', # Assuming 'starts' is a DateTimeField/DateField
            #'End Date': 'ends',     # Assuming 'ends' is a DateTimeField/DateField
         
            #'Status': 'status', # Example status field
        },
        #'related_field_lookups': {
            # No related fields for CommonCurrencyModel in this example
        #},
        #'default_values': {
            #'is_active': True, # Default value if 'Is Active' column is missing or empty
            #'status': 'active', # Default status if 'Status' column is missing or empty
        #},
    },
    
    
    
    
    'CommonStocksModel': {
        'model': 'CommonStocksModel',
        'required_excel_headers': ['Part Number', 'Description', 'Unit Cost', 'Unit Price', 'Quantity'],
        'field_mapping': {
            'Part Number': 'partNumber',
            'Description': 'description',
            'Unit Cost': 'unitcost',
            'Unit Price': 'unitPrice',
            'Quantity': 'quantity',
            'Category': 'category', # Excel column 'Category' maps to 'category' ForeignKey field
            'Unit of Measure': 'unit_of_measure', # Excel column 'Unit of Measure' maps to 'unit_of_measure' ForeignKey field
            'Inventory Account': 'inventory_account', # Excel column 'Inventory Account' maps to 'inventory_account' ForeignKey field
            'Expense Account': 'expense_account',
            'Income Account': 'income_account',
            'Tax Rate Name': 'taxrate', # Assuming taxrate is a ForeignKey to CommonTaxParametersModel
        },
        'related_field_lookups': {
            'category': {'model': 'CommonCategoriesModel', 'lookup_field': 'name'}, # Lookup Category by its 'name'
            'unit_of_measure': {'model': 'CommonUnitsOfMeasureModel', 'lookup_field': 'name'}, # Lookup UnitOfMeasure by 'name'
            'inventory_account': {'model': 'CommonChartofAccountsModel', 'lookup_field': 'name'}, # Lookup ChartOfAccounts by 'name'
            'expense_account': {'model': 'CommonChartofAccountsModel', 'lookup_field': 'name'},
            'income_account': {'model': 'CommonChartofAccountsModel', 'lookup_field': 'name'},
            'taxrate': {'model': 'CommonTaxParametersModel', 'lookup_field': 'name'}, # Lookup TaxRate by 'name'
        },
        'default_values': {
            'status': 'active',
            'is_active': True,
            'buyingPrice': Decimal('0.00'), # Default if not provided
            'standardUnitCost': Decimal('0.00'),
            'total_units_sold': 0,
        },
    },
    # Add more configurations for other models here (e.g., CommonCustomersModel, CommonSuppliersModel)
    # Example for CommonCustomersModel
    'CommonCustomersModel': {
        'model': 'CommonCustomersModel',
        'required_excel_headers': ['Customer Name', 'Email'],
        'field_mapping': {
            'Customer Name': 'name',
            'Email': 'email',
            'Phone': 'phone',
            'Address': 'address',
            'Balance': 'balance',
            'Turnover': 'turnover',
            'Payment Term': 'payment_term',
        },
        'related_field_lookups': {
            'payment_term': {'model': 'CommonPaymentTermModel', 'lookup_field': 'name'},
        },
        'default_values': {
            'balance': Decimal('0.00'),
            'turnover': Decimal('0.00'),
            'status': 'active',
            'is_active': True,
        },
    },
    # Example for CommonSuppliersModel (similar to customers)
    'CommonSuppliersModel': {
        'model': 'CommonSuppliersModel',
        'required_excel_headers': ['Supplier Name', 'Email'],
        'field_mapping': {
            'Supplier Name': 'name',
            'Email': 'email',
            'Phone': 'phone',
            'Address': 'address',
            'Balance': 'balance',
            'Turnover': 'turnover',
            'Payment Term': 'payment_term',
        },
        'related_field_lookups': {
            'payment_term': {'model': 'CommonPaymentTermModel', 'lookup_field': 'name'},
        },
        'default_values': {
            'balance': Decimal('0.00'),
            'turnover': Decimal('0.00'),
            'status': 'active',
            'is_active': True,
        },
    },
}



allif_sector_redirect_map= {
    "Sales": {
        "home": "allifmaalsalesapp:salesHome",
        "dashboard": "allifmaalsalesapp:salesDashboard",
    },
    "Healthcare": {
        "home": "allifmaalshaafiapp:shaafiHome",
        "dashboard": "allifmaalshaafiapp:shaafiDashboard",
    },
    "Hospitality": {
        "home": "allifmaalhotelsapp:hotelsHome",
        "dashboard": "allifmaalhotelsapp:hospitalityDashboard", # Note: your original had hospitalityDashboard here
    },
    "Education": {
        "home": "allifmaalilmapp:ilmHome",
        "dashboard": "allifmaalilmapp:ilmDashboard",
    },
    "Service": {
        "home": "allifmaalservicesapp:servicesHome",
        "dashboard": "allifmaalservicesapp:servicesDashboard",
    },
    "Realestate": {
        "home": "allifmaalrealestateapp:realestateHome",
        "dashboard": "allifmaalrealestateapp:realestateDashboard",
    },
    "Logistics": {
        "home": "allifmaallogisticsapp:logisticsHome",
        "dashboard": "allifmaallogisticsapp:logisticsDashboard",
    },
    # Add other sectors as needed
}

def allif_filtered_and_sorted_queryset(request: HttpRequest,model_class: type[Model], allif_data: dict,explicit_scope: Optional[str] = None) -> QuerySet:
    """
    Applies common filtering (company, scope, access levels) and sorting to a queryset.
    Retrieves sort_mapping and default_sort_field from MODEL_SORT_CONFIGS.

    Args:
        request (HttpRequest): The HttpRequest object.
        model_class (type[models.Model]): The Django Model class to query (e.g., CommonExpensesModel).
        allif_data (dict): Dictionary containing user access levels and company ID.
        explicit_scope (Optional[str]): If provided ('active', 'all', 'archived'),
                                        this will override the 'scope' from request.GET.

    Returns:
        QuerySet: The filtered and sorted QuerySet.
    """
    company_id = allif_data.get("main_sbscrbr_entity")

    model_identifier = model_class._meta.model_name 
    sort_config = allif_model_sort_configs.get(model_identifier, {})
    sort_mapping = sort_config.get('sort_mapping', {})
    default_sort_field = sort_config.get('default_sort_field', '-starts')
    default_ui_label = sort_config.get('default_ui_label', 'Default Sort')
    
    """
    do the data filtering here based on the company and other parameters like divisions, branches departments etc.
    """
    # --- Determine the effective scope ---
    # Prioritize explicit_scope, then request.GET.get('scope'), then default to 'active'
    scope = explicit_scope if explicit_scope is not None else request.GET.get('archived', 'active') 
    # above... if explicit_scope is specified, then use it, else, show all both active and archived data
    
    # Initialize queryset based on scope
    if scope == 'all' and hasattr(model_class, 'all_objects'):
        queryset = model_class.all_objects.all()
    elif scope == 'archived' and hasattr(model_class.objects, 'archived'):
        queryset = model_class.objects.archived()
    else: # Default to 'active' or standard manager
        queryset = model_class.objects.all() 

    # --- Apply company filter ---
    if hasattr(model_class, 'company') and company_id:
        queryset = queryset.filter(company=company_id)
    else:
        pass 
    
    # --- Apply access level filtering (division, branch, department) ---
    if allif_data.get("logged_in_user_has_universal_access"):
        pass 
    elif allif_data.get("logged_in_user_has_divisional_access"):
        if hasattr(model_class, 'division'):
            queryset = queryset.filter(division=allif_data.get("logged_user_division"))
        else:
            queryset = model_class.objects.none()
    elif allif_data.get("logged_in_user_has_branches_access"):
        if hasattr(model_class, 'division') and hasattr(model_class, 'branch'):
            queryset = queryset.filter(
                division=allif_data.get("logged_user_division"),
                branch=allif_data.get("logged_user_branch")
            )
        else:
            queryset = model_class.objects.none()
    elif allif_data.get("logged_in_user_has_departmental_access"):
        if hasattr(model_class, 'division') and hasattr(model_class, 'branch') and hasattr(model_class, 'department'):
            queryset = queryset.filter(
                division=allif_data.get("logged_user_division"),
                branch=allif_data.get("logged_user_branch"),
                department=allif_data.get("logged_user_department")
            )
        else:
            queryset = model_class.objects.none()
    else:
        queryset = model_class.objects.none()

    # --- Determine and apply sorting based on POST data (from your form) ---
    selected_ui_label = request.POST.get('sort_option') 
    
    if not selected_ui_label: 
        selected_ui_label = request.GET.get('sort_ui_label')

    sort_field = sort_mapping.get(selected_ui_label, default_sort_field)
    
    if not selected_ui_label or selected_ui_label not in sort_mapping:
        selected_ui_label = default_ui_label

    queryset = queryset.order_by(sort_field)

    queryset.current_sort_ui_label = selected_ui_label
    queryset.sort_options = sort_mapping.items() 
    queryset.current_scope = scope # This will reflect the *actual* scope used (explicit or from GET)

    return queryset



# --- NEW UTILITY FUNCTION FOR FORM QUERYSET INITIALIZATION ---
def allif_initialize_form_select_querysets(form_instance: forms.Form, allifmaalparameter: str, field_model_map: dict):
    """
    Initializes querysets for Select/ModelChoiceFields in a form based on a parameter.
    ...Explanation of allif_initialize_form_select_querysets:

    form_instance: The actual form object being initialized (self from your __init__ method).

    allifmaalparameter: Your allifmaalparameter (likely the company ID).

    field_model_map: This is the key. It's a dictionary that tells the function which form field corresponds to which Django Model.

    Example: {'items': CommonStocksModel, 'trans_number': CommonTransactionsModel}

    Logic:

    If allifmaalparameter exists, it iterates through the field_model_map and sets the queryset for each field to Model.objects.filter(company=allifmaalparameter).

    If allifmaalparameter does not exist, it sets the queryset for each field to Model.objects.none().

    It includes checks (if field_name in form_instance.fields and isinstance(...)) to make it more robust, ensuring it only tries to set querysets on valid ModelChoiceFields that actually exist in the form.

    Args:
        form_instance (forms.Form): The instance of the Django form.
        allifmaalparameter (str): The parameter (e.g., company ID) to filter by.
        field_model_map (dict): A dictionary where keys are field names (strings)
                                and values are the Django Model classes (e.g., CommonStocksModel).
                                Example: {'items': CommonStocksModel, 'trans_number': CommonTransactionsModel}
    """
    if allifmaalparameter:
        for field_name, model_class in field_model_map.items():
            if field_name in form_instance.fields:
                # Check if the field is indeed a ModelChoiceField or similar
                if isinstance(form_instance.fields[field_name], (forms.ModelChoiceField, forms.ModelMultipleChoiceField)):
                    form_instance.fields[field_name].queryset = model_class.objects.filter(company=allifmaalparameter)
                else:
                    pass
                    #print(f"WARNING: Field '{field_name}' in form is not a ModelChoiceField/ModelMultipleChoiceField. Skipping queryset initialization.")
            else:
                pass
                #print(f'WARNING: Field '{field_name}' not found in form fields. Skipping queryset initialization.')
    else:
        for field_name, model_class in field_model_map.items():
            if field_name in form_instance.fields:
                if isinstance(form_instance.fields[field_name], (forms.ModelChoiceField, forms.ModelMultipleChoiceField)):
                    form_instance.fields[field_name].queryset = model_class.objects.none()
            # No need for else here, if field not in form_instance.fields, it's already skipped


# --- NEW: Generic Form Submission and Save Logic Function ---
#@logged_in_user_must_have_profile
#@logged_in_user_can_view
def allif_common_form_submission_and_save____checkwithotherfunctionthendelete(request,form_class: type[forms.ModelForm],title_text: str, 
    success_redirect_url_name: str, # for the redirection url
    template_path: str,# function specific template
    
    # New optional parameter for custom pre-save logic
    pre_save_callback: Optional[callable] = None, # argument. This callback will be a function that
    # the calling view provides to perform any model-specific assignments before the object is saved.
    
    
    # New optional parameter for initial data (for GET request forms)
    initial_data: Optional[dict] = None,
    # New optional parameter for extra form arguments (for form __init__)
    extra_form_args: Optional[list] = None,
    extra_context: Optional[dict] = None,
    
    # NEW OPTIONAL PARAMETER FOR REDIRECTION LOGIC
    redirect_with_pk: bool = False, # Set to True if the success_redirect_url_name expects a 'pk' argument
    # --- NEW ADDITION START: Parameter for the PK value to use in redirect ---
    redirect_pk_value: Optional[int] = None, # <--- NEW: The specific PK to use for redirection if redirect_with_pk is True
    # --- NEW ADDITION END ---
    
    ):
    
    """
    Helper function to encapsulate the common logic for processing form submissions
    and saving new items, including custom pre-save assignments.

    Args:
        request (HttpRequest): The HttpRequest object.
        form_class (type[forms.ModelForm]): The specific ModelForm class to use.
        title_text (str): The title for the page.
        success_redirect_url_name (str): The Django URL name to redirect to on successful form submission.
        template_path (str): The path to the template to render.
        pre_save_callback (callable, optional): A function (obj, request, allif_data) -> None
                                                that performs model-specific assignments before obj.save().
        initial_data (dict, optional): Initial data for the form on GET request.
        extra_form_args (list, optional): Additional positional arguments to pass to form_class.__init__.
    """
    allif_data = common_shared_data(request)
    company_id = allif_data.get("main_sbscrbr_entity").id
    user_var = allif_data.get("usrslg")
    glblslug = allif_data.get("compslg")
    print(f"DEBUG UTILS: 4. allif_common_form_submission_and_save received extra_context keys: {(extra_context or {}).keys()}")

    # Prepare form arguments for __init__
    form_args = [company_id] # Default first argument for your forms
    if extra_form_args:
        form_args.extend(extra_form_args)

    if request.method == 'POST':
        form = form_class(*form_args, request.POST) 
        if form.is_valid():
            obj = form.save(commit=False)
            
            # --- Assign common fields from allif_data (fetching FK instances) ---
            # These fields are expected to be on CommonBaseModel and inherited by 'obj'
            
            # Company
            if hasattr(obj, 'company') and company_id:
                try:
                    obj.company = get_object_or_404(CommonCompanyDetailsModel, pk=company_id)
                except Http404: # More specific exception for get_object_or_404
                    #print(f"WARNING: Company with ID {company_id} not found for {obj.__class__.__name__}.")
                    obj.company = None 
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve company with ID {company_id}: {e}")
                    obj.company = None
            else:
                pass
            # Division
            if hasattr(obj, 'division') and allif_data.get("logged_user_division").id:
                try:
                    obj.division = get_object_or_404(CommonDivisionsModel, pk=allif_data.get("logged_user_division").id)
                except Http404:
                    #print(f"WARNING: Division with ID {allif_data.get('logged_user_division').id} not found for {obj.__class__.__name__}.")
                    obj.division = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve division with ID {allif_data.get('logged_user_division').id}: {e}")
                    obj.division = None
            else:
                pass
            # Branch
            if hasattr(obj, 'branch') and allif_data.get("logged_user_branch").id:
                try:
                    obj.branch = get_object_or_404(CommonBranchesModel, pk=allif_data.get("logged_user_branch").id)
                except Http404:
                    #print(f"WARNING: Branch with ID {allif_data.get('logged_user_branch').id} not found for {obj.__class__.__name__}.")
                    obj.branch = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve branch with ID {allif_data.get('logged_user_branch').id}: {e}")
                    obj.branch = None
            else:
                pass
            # Department
            if hasattr(obj, 'department') and allif_data.get("logged_user_department").id:
                try:
                    obj.department = get_object_or_404(CommonDepartmentsModel, pk=allif_data.get("logged_user_department").id)
                except Http404:
                    #print(f"WARNING: Department with ID {allif_data.get('logged_user_department').id} not found for {obj.__class__.__name__}.")
                    obj.department = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve department with ID {allif_data.get('logged_user_department').id}: {e}")
                    obj.department = None
            if allif_data.get("logged_user_operation_year"):
                if hasattr(obj, 'operation_year') and allif_data.get("logged_user_operation_year").id:
                    try:
                        
                        obj.operation_year = get_object_or_404(CommonOperationYearsModel, pk=allif_data.get("logged_user_operation_year").id)
                    except Http404:
                        #print(f"WARNING: Operation year with ID {allif_data.get('logged_user_operation_year').id} not found for {obj.__class__.__name__}.")
                        obj.operation_year = None
                        raise Http404("Unauthorized: Item does not belong to your company or access denied.")
                    except Exception as e:
                        #print(f"ERROR: Failed to retrieve operation year with ID {allif_data.get('logged_user_operation_year').id}: {e}")
                        obj.operation_year = None
                else:
                    pass
            else:
                pass
            if allif_data.get("logged_user_operation_term"):
                if hasattr(obj, 'operation_term') and allif_data.get("logged_user_operation_term").id:
                    try:
                        
                        obj.operation_term = get_object_or_404(CommonOperationYearTermsModel, pk=allif_data.get("logged_user_operation_term").id)
                    except Http404:
                        #print(f"WARNING: Operation term with ID {allif_data.get('logged_user_operation_term').id} not found for {obj.__class__.__name__}.")
                        obj.operation_term = None
                    except Exception as e:
                        #print(f"ERROR: Failed to retrieve operation term with ID {allif_data.get('logged_user_operation_term').id}: {e}")
                        obj.operation_term = None
                else:
                    pass
            else:
                pass
                    
            # Owner (assuming 'usernmeslg' in allif_data is the User object)
            if hasattr(obj, 'owner') and allif_data.get("usernmeslg"):
                try:
                    obj.owner=allif_data.get("usernmeslg") # This should be the User object
                except Http404:
                    #print(f"WARNING: User {allif_data.get("usernmeslg")} not found for {obj.__class__.__name__}.")
                    obj.owner = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve User {allif_data.get("usernmeslg")}: {e}")
                    obj.owner = None
            else:
                pass
            # --- Execute custom pre-save callback if provided ---
            if pre_save_callback:
                try:
                    pre_save_callback(obj, request, allif_data)
                except Exception as e:
                    #print(f"ERROR: Pre-save callback failed for {obj.__class__.__name__}: {e}")
                    # You might want to add a user-facing error message here
                    form.add_error(None, f"An internal error occurred during custom processing: {e}")
                    # Re-render the form with errors
                    context = {
                        "form": form,
                        "title": title_text,
                        "user_var": allif_data.get("usrslg"),
                        "glblslug": allif_data.get("compslg"),
                        "error_message": form.errors, # Pass form errors
                    }
                    return render(request, template_path, context)

            else:
                pass
            obj.save() # Finally save the object after all assignments
            
            # --- NEW ADDITION START: Dynamic Redirection Logic (using redirect_pk_value) ---
            # Initialize kwargs with the standard user and company slugs
            redirect_kwargs = {'allifusr': user_var, 'allifslug': glblslug}
            
            # Conditionally add 'pk' to kwargs ONLY if redirect_with_pk is True
            # AND if a redirect_pk_value has been provided.
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs['pk'] = redirect_pk_value # Use the explicitly provided PK for redirection
            elif redirect_with_pk and redirect_pk_value is None:
                # This is a warning/error case: redirect_with_pk is True but no value was given.
                # Fallback to obj.pk, but log a warning.
                logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to obj.pk ({obj.pk}).")
                redirect_kwargs['pk'] = obj.pk
            
            try:
                # Attempt to reverse the URL with the dynamically built kwargs
                final_redirect_url = reverse(f'allifmaalcommonapp:{success_redirect_url_name}', kwargs=redirect_kwargs)
            except NoReverseMatch as e:
                # If reverse fails (e.g., due to missing/extra arguments for the URL pattern),
                # log the error and provide a fallback redirect.
                logger.error(f"Failed to reverse URL '{success_redirect_url_name}' with kwargs {redirect_kwargs}: {e}")
                messages.error(request, f"Successfully created, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{success_redirect_url_name}'.")
                # Fallback to a generic home page if redirection fails
                return redirect(reverse('allifmaalcommonapp:commonHome', kwargs={'allifusr': user_var, 'allifslug': glblslug}))

            return redirect(final_redirect_url)
            # --- NEW ADDITION END ---
            
            # Redirect to the specified list URL with user and company slugs....previous...........
            #return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}',kwargs={'allifusr': allif_data.get("usrslg"), 'allifslug': allif_data.get("compslg")}))
        else:
            # Form is invalid, render with errors
            error_message = form.errors
            allifcontext = {"error_message": error_message, "title": title_text}
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    else:
        # GET request, initialize empty form
        form = form_class(*form_args, initial=initial_data)

    context = {
        "form": form,
        "title": title_text,
        "user_var": allif_data.get("usrslg"), 
        "glblslug": allif_data.get("compslg"), 
    }
    return render(request, template_path, context)



# --- NEW: Generic Edit Item LOGIC Function ---
def allif_common_form_edit_and_save(request,pk: int,form_class: type[forms.ModelForm],title_text: str, 
    success_redirect_url_name: str, 
    template_path: str,
    is_edit_mode: bool = True,
    pre_save_callback: Optional[callable] = None,
    extra_form_args: Optional[list] = None,
    extra_context: Optional[dict] = None, # For passing additional context to template
    
    # NEW OPTIONAL PARAMETER FOR REDIRECTION LOGIC
    redirect_with_pk: bool = False, # Set to True if the success_redirect_url_name expects a 'pk' argument
    # --- NEW ADDITION START: Parameter for the PK value to use in redirect ---
    redirect_pk_value: Optional[int] = None, # <--- NEW: The specific PK to use for redirection if redirect_with_pk is True
    # --- NEW ADDITION END ---
     # --- NEW ADDITION START: Parameter for the app namespace ---
    app_namespace: Optional[str] = 'allifmaalcommonapp', # <--- NEW: The app namespace for redirection URL
    # --- NEW ADDITION END ---
    
):
    """
    Helper function to encapsulate the common logic for processing form submissions
    and saving updates to existing items.

    Args:
        request (HttpRequest): The HttpRequest object.
        pk (int): Primary key of the object to be edited.
        form_class (type[forms.ModelForm]): The specific ModelForm class to use.
        title_text (str): The title for the page.
        success_redirect_url_name (str): The Django URL name to redirect to on successful form submission.
        template_path (str): The path to the template to render.
        pre_save_callback (callable, optional): A function (obj, request, allif_data) -> None
                                                that performs model-specific assignments before obj.save().
        extra_form_args (list, optional): Additional positional arguments to pass to form_class.__init__.
        extra_context (dict, optional): Additional context variables to pass to the template.
    """
    allif_data = common_shared_data(request)
    company_id = allif_data.get("main_sbscrbr_entity").id
    user_var = allif_data.get("usrslg")
    glblslug = allif_data.get("compslg")
    ahmedyaremuse="am tesint this gy"
    #logged_user_div=request.user.userdivision
    # Determine the model class from the form's Meta
    model_class = form_class.Meta.model
    if not model_class:
        raise ValueError(f'Form {form_class.__name__} does not have a model defined in its Meta class.')

    # Retrieve the object to be updated using all_objects to bypass default managers
    # and get_object_or_404 for robust error handling.
    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
   
    # --- Authorization Check (Crucial for multi-tenant/access control) ---
    # This is a placeholder. Implement your actual granular authorization logic here.
    if hasattr(allifquery, 'company')==100000:# you can remove this condition... just place  holder...
        raise Http404("Unauthorized: Item does not belong to your company or access denied.")
   
    # Add more granular checks (division, branch, department) if necessary,
    # similar to what you have in get_filtered_and_sorted_queryset.
   
    # Prepare form arguments for __init__ (company_id is typically the first)
    form_args = [company_id] 
    if extra_form_args:
        form_args.extend(extra_form_args)

    if request.method == 'POST':
        form = form_class(*form_args, request.POST, instance=allifquery) 
        if form.is_valid():
            obj = form.save(commit=False) # obj is now item_instance with updated data
            if hasattr(obj, 'division') and allif_data.get("logged_user_division").id:
            #if hasattr(obj, 'division'):
                print(allif_data.get("logged_user_division").id,'kkkkkkkkkkkkkkkkkk')
                try:
                    #obj.division = get_object_or_404(CommonDivisionsModel, pk=allif_data.get("logged_user_division").id)
                    obj.division =allif_data.get("logged_user_division")
                except Http404:
                    #print(f"WARNING: Division with ID {allif_data.get('logged_user_division').id} not found for {obj.__class__.__name__}.")
                    obj.division = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve division with ID {allif_data.get('logged_user_division').id}: {e}")
                    obj.division = None
            else:
                pass
            # Branch
            if hasattr(obj, 'branch'):
                try:
                    obj.branch =allif_data.get("logged_user_branch")
                except Http404:
                    #print(f"WARNING: Branch with ID {allif_data.get('logged_user_branch').id} not found for {obj.__class__.__name__}.")
                    obj.branch = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve branch with ID {allif_data.get('logged_user_branch').id}: {e}")
                    obj.branch = None
            else:
                pass
            # Department
            if hasattr(obj, 'department'):
                try:
                    obj.department =allif_data.get("logged_user_department")
                except Http404:
                    #print(f"WARNING: Department with ID {allif_data.get('logged_user_department').id} not found for {obj.__class__.__name__}.")
                    obj.department = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve department with ID {allif_data.get('logged_user_department').id}: {e}")
                    obj.department = None
            if allif_data.get("logged_user_operation_year"):
                if hasattr(obj, 'operation_year') and allif_data.get("logged_user_operation_year").id:
                    try:
                        
                        obj.operation_year =allif_data.get("logged_user_operation_year")
                    except Http404:
                        #print(f"WARNING: Operation year with ID {allif_data.get('logged_user_operation_year').id} not found for {obj.__class__.__name__}.")
                        obj.operation_year = None
                        raise Http404("Unauthorized: Item does not belong to your company or access denied.")
                    except Exception as e:
                        #print(f"ERROR: Failed to retrieve operation year with ID {allif_data.get('logged_user_operation_year').id}: {e}")
                        obj.operation_year = None
                else:
                    pass
            else:
                pass
            if allif_data.get("logged_user_operation_term"):
                if hasattr(obj, 'operation_term') and allif_data.get("logged_user_operation_term").id:
                    try:
                        
                        obj.operation_term =allif_data.get("logged_user_operation_term")
                    except Http404:
                        #print(f"WARNING: Operation term with ID {allif_data.get('logged_user_operation_term').id} not found for {obj.__class__.__name__}.")
                        obj.operation_term = None
                    except Exception as e:
                        #print(f"ERROR: Failed to retrieve operation term with ID {allif_data.get('logged_user_operation_term').id}: {e}")
                        obj.operation_term = None
                else:
                    pass
            else:
                pass
                    
            # Owner (assuming 'usernmeslg' in allif_data is the User object)
            if hasattr(obj, 'owner') and allif_data.get("usernmeslg"):
                try:
                    obj.owner=allif_data.get("usernmeslg") # This should be the User object
                except Http404:
                    #print(f"WARNING: User {allif_data.get("usernmeslg")} not found for {obj.__class__.__name__}.")
                    obj.owner = None
                except Exception as e:
                    #print(f"ERROR: Failed to retrieve User {allif_data.get("usernmeslg")}: {e}")
                    obj.owner = None
            else:
                pass
            

            # --- Assign common 'updated_by' field ---
            #if hasattr(obj, 'updated_by') and allif_data.get("usernmeslg"):
                #obj.updated_by = allif_data.get("usernmeslg") # This should be the User object
                

            # --- Execute custom pre-save callback if provided ---
            if pre_save_callback:
                try:
                    pre_save_callback(obj, request, allif_data)
                except Exception as e:
                    print(f"ERROR: Pre-save callback failed for {obj.__class__.__name__} (edit): {e}")
                    form.add_error(None, f"An internal error occurred during custom processing: {e}")
                    context = {
                        "form": form, "title": title_text, "allifquery": allifquery,
                        "user_var": allif_data.get("usrslg"), "glblslug": allif_data.get("compslg"),
                        "error_message": form.errors,
                        **(extra_context or {}),
                        "ahmedyaremuse":ahmedyaremuse,
                    }
                    return render(request, template_path, context)
            
           
            obj.save() # Save the updated object
            # --- NEW ADDITION START: Dynamic Redirection Logic (using app_namespace) ---..
            redirect_kwargs = {'allifusr': user_var, 'allifslug': glblslug}
            
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs['pk'] = redirect_pk_value
            elif redirect_with_pk and redirect_pk_value is None:
                # For edit/detail, if redirect_pk_value isn't provided, use the current object's PK
                logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to current item's PK ({obj.pk}).")
                redirect_kwargs['pk'] = obj.pk
            
            try:
                # Construct the full URL name using the app_namespace
                full_url_name = f'{app_namespace}:{success_redirect_url_name}'
                final_redirect_url = reverse(full_url_name, kwargs=redirect_kwargs)
            except NoReverseMatch as e:
                logger.error(f"Failed to reverse URL '{full_url_name}' with kwargs {redirect_kwargs}: {e}")
                messages.error(request, f"Successfully saved, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{full_url_name}'.")
                # Fallback to a generic home page if redirection fails
                return redirect(reverse(f'allifmaalcommonapp:commonHome', kwargs={'allifusr': user_var, 'allifslug': glblslug}))

            return redirect(final_redirect_url)
            # --- NEW ADDITION END ---
            
            """
            messages.success(request, f"{title_text} created successfully!")
            redirect_kwargs = {'allifusr': user_var, 'allifslug': glblslug}
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs['pk'] = redirect_pk_value
            elif redirect_with_pk and redirect_pk_value is None:
                logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to obj.pk ({obj.pk}).")
                redirect_kwargs['pk'] = obj.pk
            try:
                final_redirect_url = reverse(f'allifmaalcommonapp:{success_redirect_url_name}', kwargs=redirect_kwargs)
            except NoReverseMatch as e:
                logger.error(f"Failed to reverse URL '{success_redirect_url_name}' with kwargs {redirect_kwargs}: {e}")
                messages.error(request, f"Successfully created, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{success_redirect_url_name}'.")
                return redirect(reverse('allifmaalcommonapp:commonHome', kwargs={'allifusr': user_var, 'allifslug': glblslug}))
            return redirect(final_redirect_url)
            """
           
            
            #return redirect(
                #reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                        #kwargs={'allifusr': allif_data.get("usrslg"), 'allifslug': allif_data.get("compslg")})
            #)
        else:
            # Form is invalid, render with errors
            error_message = form.errors
            allifcontext = {
                "error_message": error_message, 
                "title": title_text, 
                "allifquery": allifquery, # Pass instance back for error rendering
                "user_var": allif_data.get("usrslg"), 
                "glblslug": allif_data.get("compslg"),
                
                  "form": form, # Pass the form with errors
                "title": title_text,
                "user_var": user_var,
                "glblslug": glblslug,
                "ahmedyaremuse":ahmedyaremuse,
                **(extra_context or {}) # Crucially, include extra_context here
            }
            allifcontext.update(extra_context or {})
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    else:
        # GET request, initialize form with instance data
        form = form_class(*form_args, instance=allifquery)

    context = {
        "form": form,
        "title": title_text,
        "allifquery":allifquery, # Pass the instance to the template for display
        "user_var": allif_data.get("usrslg"), 
        "glblslug": allif_data.get("compslg"), 
        **(extra_context or {}),
        "ahmedyaremuse":ahmedyaremuse,
        
     "form": form,
        "title": title_text,
        "user_var": user_var, 
        "glblslug": glblslug, 
        **(extra_context or {}) # Include extra_context here
    }
    return render(request, template_path, context)

#.... before de

# C:\am\allifapp\allifapperp\allifmaalcommonapp\utils.py

# ... (Your existing imports) ...
from django.urls import reverse, NoReverseMatch # Make sure this is imported
from typing import Optional, Callable # Make sure these are imported

# ... (Your existing model imports, logger, common_shared_data, etc.) ...

def allif_common_form_submission_and_save(
    request,
    form_class: type[forms.ModelForm],
    title_text: str,
    success_redirect_url_name: str, # for the redirection url
    template_path: str,# function specific template
    
    pre_save_callback: Optional[Callable] = None,
    initial_data: Optional[dict] = None,
    extra_form_args: Optional[list] = None,
    
    extra_context: Optional[dict] = None, # For passing additional context to template
    redirect_with_pk: bool = False, # Set to True if the success_redirect_url_name expects a 'pk' argument
    redirect_pk_value: Optional[int] = None, # The specific PK to use for redirection if redirect_with_pk is True
     # --- NEW ADDITION START: Parameter for the app namespace ---
    app_namespace: Optional[str] = 'allifmaalcommonapp', # <--- NEW: The app namespace for redirection URL
    # --- NEW ADDITION END ---
):
    """
    Helper function to encapsulate the common logic for processing form submissions
    and saving new items, including custom pre-save assignments.
    Dynamically handles redirection URLs with or without a 'pk' argument
    based on the 'redirect_with_pk' flag.
    Ensures that extra_context is always passed to the template, even on form validation errors.
    """
    allif_data = common_shared_data(request)
    company_id = allif_data.get("main_sbscrbr_entity").id
    user_var = allif_data.get("usrslg")
    glblslug = allif_data.get("compslg")

    print(f"DEBUG UTILS: 4. allif_common_form_submission_and_save received extra_context keys: {(extra_context or {}).keys()}")

    # Prepare form arguments for __init__
    form_args = [company_id] # Default first argument for your forms
    if extra_form_args:
        form_args.extend(extra_form_args)

    if request.method == 'POST':
        form = form_class(*form_args, request.POST) 
        if form.is_valid():
            obj = form.save(commit=False)
            
            # --- Assign common fields from allif_data (fetching FK instances) ---
            if hasattr(obj, 'company') and allif_data.get("main_sbscrbr_entity"):
                obj.company = allif_data.get("main_sbscrbr_entity")
            if hasattr(obj, 'division') and allif_data.get("logged_user_division"):
                obj.division = allif_data.get("logged_user_division")
            if hasattr(obj, 'branch') and allif_data.get("logged_user_branch"):
                obj.branch = allif_data.get("logged_user_branch")
            if hasattr(obj, 'department') and allif_data.get("logged_user_department"):
                obj.department = allif_data.get("logged_user_department")
            if allif_data.get("logged_user_operation_year"):
                if hasattr(obj, 'operation_year') and allif_data.get("logged_user_operation_year").id:
                    try:
                        obj.operation_year = get_object_or_404(CommonOperationYearsModel, pk=allif_data.get("logged_user_operation_year").id)
                    except Http404:
                        obj.operation_year = None
                    except Exception as e:
                        logger.error(f"ERROR: Failed to retrieve operation year with ID {allif_data.get('logged_user_operation_year').id}: {e}")
                        obj.operation_year = None
                else:
                    pass
            else:
                pass
            if allif_data.get("logged_user_operation_term"):
                if hasattr(obj, 'operation_term') and allif_data.get("logged_user_operation_term").id:
                    try:
                        obj.operation_term = get_object_or_404(CommonOperationYearTermsModel, pk=allif_data.get("logged_user_operation_term").id)
                    except Http404:
                        obj.operation_term = None
                    except Exception as e:
                        logger.error(f"ERROR: Failed to retrieve operation term with ID {allif_data.get('logged_user_operation_term').id}: {e}")
                        obj.operation_term = None
                else:
                    pass
            else:
                pass
            
            if hasattr(obj, 'owner') and allif_data.get("owner_user_object"):
                obj.owner = allif_data.get("owner_user_object")
            if hasattr(obj, 'updated_by') and request.user.is_authenticated:
                obj.updated_by = request.user
            
            if pre_save_callback:
                try:
                    pre_save_callback(obj, request, allif_data)
                except Exception as e:
                    logger.error(f"Pre-save callback failed for {obj.__class__.__name__}: {e}")
                    form.add_error(None, f"An internal error occurred during custom processing: {e}")
                    
                    # --- NEW ADDITION START: Context for pre_save_callback error (renders template_path) ---
                    context = {
                        "form": form,
                        "title": title_text,
                        "user_var": user_var,
                        "glblslug": glblslug,
                        # No "error_message" key needed here, form.errors will be used directly by template
                        **(extra_context or {}) 
                    }
                    
                    return render(request, template_path, context) # Render original template_path
                    # --- NEW ADDITION END ---

            obj.save()
            
            # --- NEW ADDITION START: Dynamic Redirection Logic (using app_namespace) ---
            redirect_kwargs = {'allifusr': user_var, 'allifslug': glblslug}
            
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs['pk'] = redirect_pk_value
            elif redirect_with_pk and redirect_pk_value is None:
                logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to obj.pk ({obj.pk}).")
                redirect_kwargs['pk'] = obj.pk
            
            try:
                # Construct the full URL name using the app_namespace
                full_url_name = f'{app_namespace}:{success_redirect_url_name}'
                final_redirect_url = reverse(full_url_name, kwargs=redirect_kwargs)
            except NoReverseMatch as e:
                logger.error(f"Failed to reverse URL '{full_url_name}' with kwargs {redirect_kwargs}: {e}")
                messages.error(request, f"Successfully created, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{full_url_name}'.")
                # Fallback to a generic home page, ensuring it also uses the app_namespace if needed
                # For commonHome, it's usually in commonapp, but being explicit is safer
                return redirect(reverse(f'allifmaalcommonapp:commonHome', kwargs={'allifusr': user_var, 'allifslug': glblslug}))

            return redirect(final_redirect_url)
            # --- NEW ADDITION END ---
            
            
            """
            messages.success(request, f"{title_text} created successfully!")
            redirect_kwargs = {'allifusr': user_var, 'allifslug': glblslug}
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs['pk'] = redirect_pk_value
            elif redirect_with_pk and redirect_pk_value is None:
                logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to obj.pk ({obj.pk}).")
                redirect_kwargs['pk'] = obj.pk
            
            try:
                final_redirect_url = reverse(f'allifmaalcommonapp:{success_redirect_url_name}', kwargs=redirect_kwargs)
            except NoReverseMatch as e:
                logger.error(f"Failed to reverse URL '{success_redirect_url_name}' with kwargs {redirect_kwargs}: {e}")
                messages.error(request, f"Successfully created, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{success_redirect_url_name}'.")
                return redirect(reverse('allifmaalcommonapp:commonHome', kwargs={'allifusr': user_var, 'allifslug': glblslug}))
            return redirect(final_redirect_url)
            """
        else:
            # Form is invalid, render with errors
            messages.error(request, "Please correct the errors below.") # Use Django messages
            
            # --- NEW ADDITION START: Context for form validation error (renders template_path) ---
            context = { # Changed from allifcontext to context for consistency
                "form": form, # Pass the form with errors
                "title": title_text,
                "user_var": user_var,
                "glblslug": glblslug,
                **(extra_context or {}) # Crucially, include extra_context here
            }
            print(f"DEBUG UTILS: 6. Context for form validation error (invalid POST): {context.keys()}")
            print(f"DEBUG UTILS:    - allifquery in invalid POST context: {context.get('allifquery')}")
            print(f"DEBUG UTILS:    - allifqueryset in invalid POST context: {context.get('allifqueryset')}")
            return render(request, template_path, context) # Render original template_path
            # --- NEW ADDITION END ---
    else:
        form = form_class(*form_args, initial=initial_data)

    # --- NEW ADDITION START: Context for GET request (renders template_path) ---
    context = {
        "form": form,
        "title": title_text,
        "user_var": user_var, 
        "glblslug": glblslug, 
        **(extra_context or {}) # Include extra_context here
    }
    print(f"DEBUG UTILS: 7. Final context for {template_path} (GET request): {context.keys()}")
    print(f"DEBUG UTILS:    - allifquery in final GET context: {context.get('allifquery')}")
    print(f"DEBUG UTILS:    - allifqueryset in final GET context: {context.get('allifqueryset')}")
    # --- NEW ADDITION END ---
    return render(request, template_path, context)




# ... (your existing ALLIF_MODEL_REGISTRY, EXCEL_UPLOAD_CONFIGS, etc.) ...
# C:\am\allifapp\allifapperp\allifmaalcommonapp\utils.py

# ... (existing imports, ensure logging, HttpRequest, HttpResponse, messages are imported) ...
from django.db.models import QuerySet # Import QuerySet for type hinting

logger = logging.getLogger(__name__)

# ... (your existing ALLIF_MODEL_REGISTRY, EXCEL_UPLOAD_CONFIGS, etc.) ...

# --- NEW: Generic Detail View Helper Function with Related Items ---
def allif_common_detail_view(
    request: HttpRequest,
    model_class, # The Django model class (e.g., CommonCategoriesModel)
    pk: int, # The primary key of the object
    template_name: str, # The path to the specific detail template
    title_map: dict = None, # Optional: For dynamic titles based on sector
    related_data_configs: list = None, # NEW: List of dicts for related data
) -> HttpResponse:
    """
    A generic function to handle detail views for various models, including related data.
    It fetches a main object by PK, determines the title, fetches related objects,
    and renders the specified template.

    Args:
        request (HttpRequest): The Django request object.
        model_class: The Django model class of the main object to fetch.
        pk (int): The primary key of the main object.
        template_name (str): The path to the specific detail template.
        title_map (dict, optional): A dictionary for dynamic titles based on sector.
                                     e.g., {"Healthcare": "Patient Details", "default": "Customer Details"}.
        related_data_configs (list, optional): A list of dictionaries, each defining
                                                how to fetch a related queryset.
                                                Example:
                                                [
                                                    {
                                                        'context_key': 'invoice_items', # Key for template context
                                                        'related_model': 'CommonInvoiceItemsModel', # String name of related model
                                                        'filter_field': 'invoice', # Field on related_model that links to main_object
                                                        'order_by': ['-date_created'], # Optional: Ordering for related items
                                                        'prefetch_related': [], # Optional: List of related fields to prefetch
                                                    },
                                                    # ... more related configs
                                                ]
    Returns:
        HttpResponse: The rendered Django template response.
    """
    try:
        allif_data = common_shared_data(request) # Get shared data early for context
        user_var = allif_data.get("usrslg")
        glblslug = allif_data.get("compslg")

        # Fetch the main object
        if hasattr(model_class, 'all_objects'):
            allifquery = model_class.all_objects.filter(pk=pk).first()
        else:
            allifquery = model_class.all_objects.filter(pk=pk).first()

        if not allifquery:
            logger.warning(f"Object of type {model_class.__name__} with PK {pk} not found.")
            messages.error(request, f"{model_class.__name__} with ID {pk} not found.")
            return render(request, 'allifmaalcommonapp/error/error.html', {'error_message': f"{model_class.__name__} not found."})

        # Determine the title dynamically
        title = str(allifquery) # Default title is the object's string representation

        if title_map:
            company_entity = allif_data.get("main_sbscrbr_entity")
            sector = None
            if company_entity and hasattr(company_entity, 'sector') and company_entity.sector:
                sector = str(company_entity.sector)

            if sector and sector in title_map:
                title = title_map[sector]
            elif 'default' in title_map:
                title = title_map['default']

        # Initialize context with main object and common data
        context = {
            "allifquery": allifquery, # The main object being detailed
            "title": title,
            "user_var": user_var,
            "glblslug": glblslug,
        }

        # Fetch related data if configurations are provided
        if related_data_configs:
            for config in related_data_configs:
                context_key = config.get('context_key')
                related_model_name = config.get('related_model')
                filter_field = config.get('filter_field')
                order_by_fields = config.get('order_by', [])
                prefetch_fields = config.get('prefetch_related', [])

                if not all([context_key, related_model_name, filter_field]):
                    logger.error(f"Incomplete related_data_config: {config}. Skipping.")
                    continue

                related_model_class = allif_main_models_registry.get(related_model_name)
                if not related_model_class:
                    logger.error(f"Related model '{related_model_name}' not found in ALLIF_MODEL_REGISTRY. Skipping.")
                    continue

                try:
                    # Construct the filter to link to the main object
                    filter_kwargs = {filter_field: allifquery}
                    
                    # Apply multi-tenancy filter if the related model has a 'company' field
                    if hasattr(related_model_class, 'company') and allif_data.get('main_sbscrbr_entity'):
                        filter_kwargs['company'] = allif_data['main_sbscrbr_entity']

                    queryset = related_model_class.all_objects.filter(**filter_kwargs)
                   

                    # Apply prefetch_related if specified
                    if prefetch_fields:
                        queryset = queryset.prefetch_related(*prefetch_fields)
                    
                    # Apply ordering if specified
                    if order_by_fields:
                        queryset = queryset.order_by(*order_by_fields)

                    context[context_key] = queryset # Add related queryset to context

                except Exception as e:
                    logger.exception(f"Error fetching related data for '{context_key}' (Model: {related_model_name}) related to {model_class.__name__} (PK: {pk}): {e}")
                    context[context_key] = [] # Provide an empty list on error to prevent template errors

        return render(request, template_name, context)

    except Exception as ex:
        logger.exception(f"Critical error in allif_common_detail_view for {model_class.__name__} (PK: {pk}): {ex}")
        error_context = {'error_message': str(ex)}
        return render(request, 'allifmaalcommonapp/error/error.html', error_context)


# --- NEW: Generic Edit Item LOGIC Function ---
def allif_delete_confirm(request,pk: int,model_name: str,title_text: str,
            template_path: str,):
    
    allif_data = common_shared_data(request)
    #model_class = model_name.model
    if not model_name:
        raise ValueError(f"Form {model_name.__name__} does not have a model defined in its Meta class.")

    allifquery = get_object_or_404(model_name.all_objects, pk=pk)
   
    context = {
        
        "title": title_text,
        "allifquery":allifquery, # Pass the instance to the template for display
        "user_var": allif_data.get("usrslg"), 
        "glblslug": allif_data.get("compslg"), 
        
    }
    return render(request, template_path, context)

def allif_delete_hanlder(request: HttpRequest,model_name: str,pk: int,success_redirect_url_name: str,
                         redirect_with_pk: bool = False,redirect_pk_value: Optional[int] = None,
                           # --- NEW ADDITION START: Parameter for the app namespace ---
    app_namespace: Optional[str] = 'allifmaalcommonapp', # <--- NEW: The app namespace for redirection URL
    # --- NEW ADDITION END ---
                         ):
    allif_data = common_shared_data(request)
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")
    model_class = allif_main_models_registry.get(model_name)
    
    
    """
    Helper function to encapsulate the common logic for deleting items.
    Dynamically handles redirection URLs with or without a 'pk' argument
    based on the 'redirect_with_pk' flag and 'redirect_pk_value'.
    Supports dynamic app namespaces for redirection URLs.
    """
    allif_data = common_shared_data(request)
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")

    model_class = allif_main_models_registry.get(model_name)
    if not model_class:
        messages.error(request, f"Error: Model '{model_name}' not found.")
        # Fallback to home if model not found
        return redirect(reverse(f'{app_namespace}:commonHome', kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    try:
        # Use .all_objects if available, otherwise .objects
        if hasattr(model_class, 'all_objects'):
            item = get_object_or_404(model_class.all_objects, pk=pk)
        else:
            item = get_object_or_404(model_class.objects, pk=pk)

        # --- NEW ADDITION START: Authorization Check (similar to edit/save) ---
        # Crucial for multi-tenant data segregation
        """if hasattr(item, 'company') and item.company != allif_data.get('main_sbscrbr_entity'):"""
        if hasattr(item, 'company') and item.company != allif_data.get('main_sbscrbr_entity'):
            logger.warning(f"Unauthorized delete attempt for {model_name} (PK: {pk}) by user {request.user} for company {allif_data.get('main_sbscrbr_entity')}.")
            messages.error(request, f"Access denied to delete {model_name} with ID {pk}.")
            # Redirect to the success URL, but without deleting
            # We need to ensure the redirect_pk_value is used here if applicable
            redirect_kwargs_unauthorized = {'allifusr': user_slug, 'allifslug': company_slug}
            if redirect_with_pk and redirect_pk_value is not None:
                redirect_kwargs_unauthorized['pk'] = redirect_pk_value
            # If no redirect_pk_value is given for a PK-requiring URL, it's an issue.
            # For delete, usually you redirect to a list or parent, so obj.pk is not relevant here.
            try:
                unauthorized_redirect_url = reverse(f'{app_namespace}:{success_redirect_url_name}', kwargs=redirect_kwargs_unauthorized)
            except NoReverseMatch as e:
                logger.error(f"Failed to reverse URL '{app_namespace}:{success_redirect_url_name}' for unauthorized delete: {e}")
                unauthorized_redirect_url = reverse(f'allifmaalcommonapp:commonHome', kwargs={'allifusr': user_slug, 'allifslug': company_slug})
            return redirect(unauthorized_redirect_url)
        # --- NEW ADDITION END ---

        item_display_name = str(item) # Get a string representation before deleting
        item.delete()
        messages.success(request, f"{model_name} '{item_display_name}' deleted successfully!")

    except Http404:
        messages.error(request, f"{model_name} with ID {pk} not found.")
    except Exception as e:
        logger.exception(f"Error deleting {model_name} (PK: {pk}): {e}")
        messages.error(request, f"An error occurred while deleting {model_name}: {e}")

    # --- NEW ADDITION START: Dynamic Redirection Logic (using app_namespace) ---
    redirect_kwargs = {'allifusr': user_slug, 'allifslug': company_slug}
    
    if redirect_with_pk and redirect_pk_value is not None:
        redirect_kwargs['pk'] = redirect_pk_value
    elif redirect_with_pk and redirect_pk_value is None:
        # For delete, if redirect_with_pk is True but no redirect_pk_value is given,
        # it usually means redirecting to a parent list or a specific parent's detail page.
        # It cannot fall back to 'obj.pk' because 'obj' (the deleted item) no longer exists.
        # This case needs careful handling in the calling view.
        logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided in delete handler. This might lead to NoReverseMatch.")
        # No automatic fallback to obj.pk here as obj is deleted.
        # The calling view MUST provide redirect_pk_value if redirect_with_pk is True.
    
    try:
        # Construct the full URL name using the app_namespace
        full_url_name = f'{app_namespace}:{success_redirect_url_name}'
        final_redirect_url = reverse(full_url_name, kwargs=redirect_kwargs)
    except NoReverseMatch as e:
        logger.error(f"Failed to reverse URL '{full_url_name}' with kwargs {redirect_kwargs}: {e}")
        messages.error(request, f"Deletion successful, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{full_url_name}'.")
        # Fallback to a generic home page if redirection fails
        return redirect(reverse(f'allifmaalcommonapp:commonHome', kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    return redirect(final_redirect_url)
    # --- NEW ADDITION END ---


"""
    if not model_class:
        raise Http404(f"Model '{model_name}' not found in mapping.")

    item = get_object_or_404(model_class.all_objects, pk=pk)

    item.delete()
    
    redirect_kwargs = {'allifusr': user_slug, 'allifslug': company_slug}
    if redirect_with_pk and redirect_pk_value is not None:
        redirect_kwargs['pk'] = redirect_pk_value
    elif redirect_with_pk and redirect_pk_value is None:
        logger.warning(f"redirect_with_pk is True for '{success_redirect_url_name}' but no redirect_pk_value was provided. Falling back to obj.pk .")
        #redirect_kwargs['pk'] = obj.pk
    
    try:
        final_redirect_url = reverse(f'allifmaalcommonapp:{success_redirect_url_name}', kwargs=redirect_kwargs)
    except NoReverseMatch as e:
        logger.error(f"Failed to reverse URL '{success_redirect_url_name}' with kwargs {redirect_kwargs}: {e}")
        messages.error(request, f"Successfully created, but failed to redirect. Please check URL configuration or 'redirect_with_pk' flag for '{success_redirect_url_name}'.")
        return redirect(reverse('allifmaalcommonapp:commonHome', kwargs={'allifusr': user_slug, 'allifslug':company_slug}))

    return redirect(final_redirect_url)
    #return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}',kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

"""

# --- NEW: Centralized Search Handler ---
def allif_search_handler(request: HttpRequest,model_name: str,search_fields_key: str, # Key to look up in SEARCH_CONFIGS
    template_path: str,
    search_input_name: str = 'allifsearchcommonfieldname', # Name of the search input field
    extra_context: Optional[dict] = None
) -> HttpRequest:
    """
    Handles generic search functionality for various models.
    """
    allif_data = common_shared_data(request)
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")
    company_id = allif_data.get("main_sbscrbr_entity")

    title = "Search Results"
    allifqueryset = []
    search_term = None

    if request.method == 'POST':
        search_term = request.POST.get(search_input_name, '').strip()
    elif request.method == 'GET': # Allow GET for initial display or direct search URLs
        search_term = request.GET.get(search_input_name, '').strip()

    if search_term:
        model_class = allif_main_models_registry.get(model_name)
        if not model_class:
            #print(f"ERROR: Search - Model '{model_name}' not found in allif_delete_models_class_map.")
            raise Http404(f"Model '{model_name}' not found for search.")

        search_fields =allif_search_config_mapping.get(search_fields_key)
        if not search_fields:
            #print(f"ERROR: Search - No search fields configured for key '{search_fields_key}' in allif_search_config_mapping.")
            raise ValueError(f"Search configuration missing for '{search_fields_key}'.")

        # Build dynamic Q objects for search
        q_objects = Q()
        for field in search_fields:
            # Assumes field names include lookup like 'description__icontains'
            q_objects |= Q(**{field: search_term})

        # Apply the search query and company filter
        queryset = model_class.objects.filter(q_objects)
        if hasattr(model_class, 'company') and company_id:
            queryset = queryset.filter(company=company_id)
        
        # Apply multi-tenancy access filters (division, branch, department)
        # This reuses logic from get_filtered_and_sorted_queryset if applicable
        if allif_data.get("logged_in_user_has_universal_access"):
            pass 
        elif allif_data.get("logged_in_user_has_divisional_access"):
            if hasattr(model_class, 'division'):
                queryset = queryset.filter(division=allif_data.get("logged_user_division"))
            else:
                queryset = model_class.objects.none()
        elif allif_data.get("logged_in_user_has_branches_access"):
            if hasattr(model_class, 'division') and hasattr(model_class, 'branch'):
                queryset = queryset.filter(
                    division=allif_data.get("logged_user_division"),
                    branch=allif_data.get("logged_user_branch")
                )
            else:
                queryset = model_class.objects.none()
        elif allif_data.get("logged_in_user_has_departmental_access"):
            if hasattr(model_class, 'division') and hasattr(model_class, 'branch') and hasattr(model_class, 'department'):
                queryset = queryset.filter(
                    division=allif_data.get("logged_user_division"),
                    branch=allif_data.get("logged_user_branch"),
                    department=allif_data.get("logged_user_department")
                )
            else:
                queryset = model_class.objects.none()
        else:
            queryset = model_class.objects.none()

        allifqueryset = queryset.distinct() # Use distinct to avoid duplicates if Q objects overlap

    context = {
        "title": title,
        "allifqueryset": allifqueryset,
        "user_var": user_slug,
        "glblslug": company_slug,
        "search_term": search_term, # Pass the search term back to the template for display
        **(extra_context or {}) # Include any extra context passed in
    }
    return render(request, template_path, context)



# --- NEW: Generic Excel Generation Utility ---
def allif_generate_excel_response(
    queryset: QuerySet, 
    fields_config: List[Dict[str, str]], # [{'field': 'model_field_name', 'label': 'Column Header'}]
    filename: str = "data.xlsx"
) -> HttpResponse:
    """
    Generates an Excel (xlsx) file from a Django QuerySet.
    'fields_config' is a list of dictionaries, where each dict has:
    - 'field': The name of the model field (can be a related field lookup like 'customer__name').
    - 'label': The desired column header in the Excel file.
    """
    print(f"DEBUG: allif_generate_excel_response called for filename: '{filename}'")
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data Export"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # Green header
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        center_aligned_text = Alignment(horizontal="center")

        # Write headers
        headers = [field_info['label'] for field_info in fields_config]
        sheet.append(headers)
        
        # Apply header styles
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_aligned_text

        # Write data rows
        for obj in queryset:
            row_data = []
            for field_info in fields_config:
                field_name = field_info['field']
                # Handle related fields (e.g., 'customer__name')
                parts = field_name.split('__')
                value = obj
                try:
                    for part in parts:
                        value = getattr(value, part)
                    # Format date/datetime objects
                    if isinstance(value, (datetime.date, datetime.datetime)):
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, datetime.datetime) else value.strftime('%Y-%m-%d')
                    elif isinstance(value, Decimal):
                         value = float(value) # Excel often prefers floats for numbers
                    elif value is None:
                        value = '' # Represent None as empty string in Excel
                    row_data.append(value)
                except AttributeError:
                    row_data.append('') # Field not found or invalid lookup
                    #print(f"WARNING: Field '{field_name}' not found on object {obj} (ID: {obj.pk}).")
                except Exception as e:
                    row_data.append(f"Error: {e}") # Catch other potential issues
                    #print(f"ERROR: Failed to get value for field '{field_name}' on object {obj} (ID: {obj.pk}): {e}")
            sheet.append(row_data)
            
        # Adjust column widths (optional)
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Create an in-memory binary stream
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0) # Rewind to the beginning of the stream

        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        print(f"DEBUG: Excel '{filename}' generated successfully.")
        return response

    except Exception as e:
        #print(f"CRITICAL ERROR: Exception during Excel generation for '{filename}': {e}")
        return HttpResponse(f'An unexpected error occurred during Excel generation: {e}', status=500)

# --- NEW: Generic PDF Generation Utility ---
def allif_generate_pdf_response(template_path: str, context: Dict[str, Any],filename: str = "document.pdf") -> HttpResponse:
    """
    Generates a PDF response from a Django template using xhtml2pdf.
    """
    template = get_template(template_path)
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{filename}"'# opens the pdf in the same page...
    response['Content-Disposition'] = f'attachment; filename="{filename}"' # downloads the pdf...
    
    try:
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            # If there's an error during PDF creation, return a generic error message
            #print(f"ERROR: PDF generation failed: {pisa_status.err}")
            return HttpResponse('We had some errors generating the PDF. Please try again.', status=500)
    except Exception as e:
        #print(f"CRITICAL ERROR: Exception during PDF generation: {e}")
        return HttpResponse(f"An unexpected error occurred during PDF generation: {e}", status=500)
    
    return response

def allif_advance_search_handler(
    request: HttpRequest,
    model_name: str,
    advanced_search_config_key: str, 
    template_html_path: str,
    template_pdf_path: Optional[str] = None, 
    extra_context: Optional[dict] = None
) -> HttpResponse:
    """
    Handles generic advanced search functionality (date/value ranges) for various models,
    including conditional PDF generation.
    """
    allif_data = common_shared_data(request)
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")
    company_id = allif_data.get("main_sbscrbr_entity")

    title = "Advanced Search Results"
    allifqueryset = [] 

    model_class = allif_main_models_registry.get(model_name) 
    if not model_class:
        #print(f"ERROR: Advanced Search - Model '{model_name}' not found in ALLIF_MODEL_REGISTRY.")
        raise Http404(f'Model "{model_name}" not found for advanced search.')

    search_config = allif_advanced_search_configs.get(advanced_search_config_key) 
    if not search_config:
        #print(f"ERROR: Advanced Search - No configuration for key '{advanced_search_config_key}' in ADVANCED_SEARCH_CONFIGS.")
        raise ValueError(f"Advanced search configuration missing for '{advanced_search_config_key}'.")

    date_field = search_config['date_field']
    value_field = search_config['value_field']
    default_order_by_date = search_config['default_order_by_date']
    default_order_by_value = search_config['default_order_by_value']
    excel_fields = search_config.get('excel_fields', []) # NEW: Get excel fields config
    # --- Calculate dynamic default values for date and amount ranges ---
    # Get the base queryset with company and access filtering
    base_queryset = allif_filtered_and_sorted_queryset(request, model_class, allif_data, explicit_scope='all') 

    # Determine earliest/latest dates and largest value from the *entire* relevant dataset
    # These will be used as defaults if user doesn't provide inputs
    first_item_by_date = base_queryset.order_by(date_field).first()
    last_item_by_date = base_queryset.order_by(f'-{date_field}').first()
    largest_item_by_value = base_queryset.order_by(default_order_by_value).first()

    # Initialize default dates as timezone-aware datetimes (midnight of current date)
    # This prevents the RuntimeWarning when comparing with DateTimeFields
    current_aware_datetime = timezone.make_aware(datetime.datetime.combine(timezone.now().date(), datetime.time.min))

    # Get default firstDate (earliest record or current_aware_datetime)
    if first_item_by_date and hasattr(first_item_by_date, date_field):
        firstDate_val = getattr(first_item_by_date, date_field)
        # If it's already a datetime, use it. If it's a date, convert to datetime.
        firstDate = timezone.make_aware(datetime.datetime.combine(firstDate_val, datetime.time.min)) if isinstance(firstDate_val, datetime.date) and not isinstance(firstDate_val, datetime.datetime) else firstDate_val
    else:
        firstDate = current_aware_datetime

    # Get default lastDate (latest record or current_aware_datetime)
    if last_item_by_date and hasattr(last_item_by_date, date_field):
        lastDate_val = getattr(last_item_by_date, date_field)
        # If it's already a datetime, use it. If it's a date, convert to datetime.
        lastDate = timezone.make_aware(datetime.datetime.combine(lastDate_val, datetime.time.max)) if isinstance(lastDate_val, datetime.date) and not isinstance(lastDate_val, datetime.datetime) else lastDate_val
    else:
        lastDate = current_aware_datetime

    largestAmount = getattr(largest_item_by_value, value_field, 0) if largest_item_by_value else 0
    largestAmount = max(0, largestAmount) 

    formats = CommonDocsFormatModel.all_objects.all()
    scopes = CommonCompanyScopeModel.all_objects.filter(company=company_id).order_by('-date')[:4] 

    # --- Process POST/GET request for search parameters ---
    if request.method == 'POST':
        selected_option = request.POST.get('requiredformat')
        start_date_str = request.POST.get('startdate')
        end_date_str = request.POST.get('enddate')
        start_value_str = request.POST.get('startvalue')
        end_value_str = request.POST.get('endvalue')
    else: 
        selected_option = request.GET.get('requiredformat')
        start_date_str = request.GET.get('startdate')
        end_date_str = request.GET.get('enddate')
        start_value_str = request.GET.get('startvalue')
        end_value_str = request.GET.get('endvalue')
    
    # Convert string inputs to appropriate types, using defaults if empty
    # Convert input dates to timezone-aware datetimes for filtering
    filter_start_date = None
    if start_date_str:
        try:
            # Parse date string to date object, then combine with min time and make aware
            filter_start_date = timezone.make_aware(datetime.datetime.combine(datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date(), datetime.time.min))
        except ValueError:
            #print(f"WARNING: Invalid start_date_str: '{start_date_str}'. Ignoring date filter.")
            filter_start_date = None

    filter_end_date = None
    if end_date_str:
        try:
            # Parse date string to date object, then combine with max time (23:59:59.999999) and make aware
            filter_end_date = timezone.make_aware(datetime.datetime.combine(datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date(), datetime.time.max))
        except ValueError:
            #print(f"WARNING: Invalid end_date_str: '{end_date_str}'. Ignoring date filter.")
            filter_end_date = None
    
    # Safely convert value strings to float, default to None if empty or invalid
    try:
        start_value = float(start_value_str) if start_value_str else None
    except ValueError:
        start_value = None
        #print(f"WARNING: Invalid start_value_str: '{start_value_str}'")
    
    try:
        end_value = float(end_value_str) if end_value_str else None
    except ValueError:
        end_value = None
        #print(f"WARNING: Invalid end_value_str: '{end_value_str}'")

    # Determine if any search criteria were provided
    search_criteria_provided = any([start_date_str, end_date_str, start_value_str, end_value_str])

    # --- Apply filters only if search criteria are provided ---
    if search_criteria_provided:
        queryset = base_queryset # Start with the base queryset
        
        # Apply date range filters using timezone-aware datetimes
        if filter_start_date:
            queryset = queryset.filter(**{f'{date_field}__gte': filter_start_date})
        else:
            queryset = queryset.filter(**{f'{date_field}__gte': firstDate}) # Use timezone-aware default
        
        if filter_end_date:
            queryset = queryset.filter(**{f'{date_field}__lte': filter_end_date})
        else:
            queryset = queryset.filter(**{f'{date_field}__lte': lastDate}) # Use timezone-aware default

        # Apply value range filters
        if start_value is not None:
            queryset = queryset.filter(**{f'{value_field}__gte': start_value})
        else:
            queryset = queryset.filter(**{f'{value_field}__gte': 0}) 
        
        if end_value is not None:
            queryset = queryset.filter(**{f'{value_field}__lte': end_value})
        else:
            queryset = queryset.filter(**{f'{value_field}__lte': largestAmount}) 
        
        allifqueryset = queryset.distinct() 
        #print(f"DEBUG: Advanced Search - Filtered results count: {searched_data.count()}")
    else:
        allifqueryset = base_queryset.distinct()
        #print(f"DEBUG: Advanced Search - No criteria provided, showing all base data. Count: {searched_data.count()}")


    # Prepare context for rendering
    context = {
        "title": title,
        "allifqueryset": allifqueryset,
        "formats": formats,
        "scopes": scopes,
        "user_var": user_slug,
        "glblslug": company_slug,
        # Pass back original string inputs for form persistence
        "start_date_input": start_date_str, 
        "end_date_input": end_date_str,
        "start_value_input": start_value_str,
        "end_value_input": end_value_str,
        # Pass calculated defaults for display if needed
        "firstDate_display": firstDate.date() if isinstance(firstDate, datetime.datetime) else firstDate, # For display only
        "lastDate_display": lastDate.date() if isinstance(lastDate, datetime.datetime) else lastDate,   # For display only
        "largestAmount_display": largestAmount,
        "selected_format_input": selected_option,
        **(extra_context or {}),
        "main_sbscrbr_entity":allif_data.get("main_sbscrbr_entity")
    }

    # --- Conditional Output (HTML or PDF) ---
    #print(f"DEBUG: Advanced Search - selected_option: '{selected_option}', template_pdf_path: '{template_pdf_path}'")
    #if selected_option == "pdf" and template_pdf_path:
        #print(f"DEBUG: Advanced Search - PDF option selected and template_pdf_path is valid. Calling allif_generate_pdf_response.")
        #return allif_generate_pdf_response( template_pdf_path,context,filename=f"{advanced_search_config_key.lower()}-advanced-search-results.pdf")
    #else:
        #print(f"DEBUG: Advanced Search - HTML option selected or template_pdf_path is invalid. Rendering HTML template: '{template_html_path}'")
        #return render(request, template_html_path, context)

     # --- Conditional Output (HTML, PDF, or Excel) ---
    if selected_option == "pdf" and template_pdf_path:
        context = {
            "title": title,
            "allifqueryset": allifqueryset, # Pass searched_data to PDF template
            "formats": formats,
            "scopes": scopes,
            "user_var": user_slug,
            "glblslug": company_slug,
            "start_date_input": start_date_str, 
            "end_date_input": end_date_str,
            "start_value_input": start_value_str,
            "end_value_input": end_value_str,
            "firstDate_display": firstDate.date() if isinstance(firstDate, datetime.datetime) else firstDate, 
            "lastDate_display": lastDate.date() if isinstance(lastDate, datetime.datetime) else lastDate,   
            "largestAmount_display": largestAmount,
            "selected_format_input": selected_option,
            **(extra_context or {}),
            "main_sbscrbr_entity":allif_data.get("main_sbscrbr_entity")
        }
        return allif_generate_pdf_response( 
            template_pdf_path,
            context,
            filename=f"{advanced_search_config_key.lower()}-advanced-search-results.pdf"
        )
    elif selected_option == "excel" and excel_fields: # NEW: Handle Excel option
        # For Excel, we directly pass the queryset and fields config
        filename = f"{advanced_search_config_key.lower()}-advanced-search-results.xlsx"
        return allif_generate_excel_response(
            allifqueryset, # Pass the queryset directly
            excel_fields,
            filename=filename
        )
    else:
        # Prepare context for HTML rendering
        context = {
            "title": title,
            "allifqueryset": allifqueryset, # Pass searched_data to HTML template
            "formats": formats,
            "scopes": scopes,
            "user_var": user_slug,
            "glblslug": company_slug,
            "start_date_input": start_date_str, 
            "end_date_input": end_date_str,
            "start_value_input": start_value_str,
            "end_value_input": end_value_str,
            "firstDate_display": firstDate.date() if isinstance(firstDate, datetime.datetime) else firstDate, 
            "lastDate_display": lastDate.date() if isinstance(lastDate, datetime.datetime) else lastDate,   
            "largestAmount_display": largestAmount,
            "selected_format_input": selected_option,
            **(extra_context or {}),
            "main_sbscrbr_entity":allif_data.get("main_sbscrbr_entity")
        }
        return render(request, template_html_path, context)



# --- Centralized Document PDF Handler ---
def allif_document_pdf_handler(request: HttpRequest,pk: int,document_config_key: str, # Key to look up in DOCUMENT_PDF_CONFIGS
    extra_context: Optional[Dict[str, Any]] = None
) -> HttpResponse:
    """
    Handles generic PDF generation for various document types (PO, Invoice, Quote, etc.).
    Fetches the main document and its related items based on configuration.
    It can generate a PDF for a main model with or without related items.
    """
    allif_data = common_shared_data(request)
    company_id = allif_data.get("main_sbscrbr_entity")
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")

    config = allif_main_document_pdf_configuration.get(document_config_key)
    if not config:
        #print(f"ERROR: PDF Handler - No configuration for key '{document_config_key}' in DOCUMENT_PDF_CONFIGS.")
        raise ValueError(f"Document PDF configuration missing for '{document_config_key}'.")

    main_model_name = config['main_model']
    # These are now optional in the config
    items_model_name = config.get('items_model') 
    items_related_field = config.get('items_related_field')
    
    title_prefix = config['title']
    filename_prefix = config['filename_prefix']
    template_path = config['template_path']
    extra_context_map = config.get('extra_context_map', {})
    related_lookups = config.get('related_lookups', [])

    main_model_class = allif_main_models_registry.get(main_model_name) 
    
    # Only try to get items_model_class if items_model_name is provided in config
    items_model_class = None
    if items_model_name:
        items_model_class = allif_main_models_registry.get(items_model_name)

    if not main_model_class: # Only main model is strictly required
        print(f"ERROR: PDF Handler - main_model_class: {main_model_class}")
        print(f"ERROR: PDF Handler - Main model class '{main_model_name}' not found in ALLIF_MODEL_REGISTRY.")
        raise Http404("Required main model not found for PDF generation.")
    
    # If items_model_name is provided but its class is not found, raise an error
    if items_model_name and not items_model_class:
        #print(f"ERROR: PDF Handler - items_model_class: {items_model_class}")
        #print(f"ERROR: PDF Handler - Items model class '{items_model_name}' not found in ALLIF_MODEL_REGISTRY, but specified in config.")
        raise Http404("Required items model not found for PDF generation, as specified in config.")

    try:
        # Fetch the main document object
        main_query = main_model_class.all_objects.filter(pk=pk)
        
        # Apply select_related/prefetch_related for optimized fetching of related data
        if related_lookups:
            main_query = main_query.select_related(*related_lookups)

        main_document = get_object_or_404(main_query, pk=pk)

        # --- Authorization Check (Crucial for multi-tenant) ---
        #if hasattr(main_document, 'company') and main_document.company and str(main_document.company.id) != company_id:
            #print(f"ERROR: Unauthorized PDF generation attempt for document {main_document.pk} from company {company_id}.")
            #raise Http404("Unauthorized: Document does not belong to your company or access denied.")

        # Fetch related items for the document ONLY IF configured
        items_queryset = None
        if items_model_class and items_related_field:
            try:
                items_queryset = items_model_class.objects.filter(**{items_related_field: main_document})
            except Exception as e:
                print(f"WARNING: Could not fetch related items for {main_model_name} (ID: {pk}): {e}")
                items_queryset = None # Ensure it's None if fetching fails

        # Prepare context for the PDF template
        context = {
            "title": title_prefix,
            "main_sbscrbr_entity": allif_data.get("main_sbscrbr_entity"),
            "scopes": CommonCompanyScopeModel.all_objects.filter(company=company_id).order_by('-date')[:4], 
            "allifquery": main_document, # The main document object
            "allifqueryset": items_queryset, # The related items (will be None if not configured/found)
            "system_user": allif_data.get("owner_user_object"), 
            "user_var": user_slug, 
            "glblslug": company_slug, 
            "timezone": timezone, 
            "main_query":main_query,
        }

        # Add any extra context fields dynamically from the main_document
        for model_field, context_name in extra_context_map.items():
            if hasattr(main_document, model_field):
                context[context_name] = getattr(main_document, model_field)

        # Merge any additional context passed from the view
        context.update(extra_context or {})

        # Generate filename
        filename = f"{filename_prefix}-{main_document.pk}.pdf" 
        if hasattr(main_document, 'name'): 
            filename = f"{filename_prefix}-{getattr(main_document, 'name')}-{main_document.pk}.pdf"
        elif hasattr(main_document, '__str__'): 
            filename = f"{filename_prefix}-{str(main_document).replace(' ', '_')}-{main_document.pk}.pdf"
        
        # Call the generic PDF generation utility
        return allif_generate_pdf_response(template_path, context, filename=filename)

    except Http404 as e:
        raise e 
    except Exception as e:
        #print(f"CRITICAL ERROR: Failed to generate PDF for {document_config_key} (ID: {pk}): {e}")
        raise Http404(f"An unexpected error occurred during PDF generation: {e}")

# --- NEW: Centralized Excel Upload Handler ---
def allif_excel_upload_handler(
    request: HttpRequest,
    model_config_key: str, # Key to look up in EXCEL_UPLOAD_CONFIGS
    success_redirect_url_name: str, # URL name to redirect to on success
) -> HttpResponse:
    """
    Handles generic Excel file uploads for various models.
    It reads an Excel file, validates data, performs type conversions,
    and creates/updates model instances in an atomic transaction.
    """
    allif_data = common_shared_data(request)
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")
    company_id = allif_data.get("main_sbscrbr_entity").id
    
    # Ensure it's a POST request and has a file
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "Please upload an Excel file.")
        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    excel_file = request.FILES['excel_file']

    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "Invalid file type. Please upload an Excel file (.xlsx or .xls).")
        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    config = allif_excel_upload_configs.get(model_config_key)
    if not config:
        messages.error(request, f"Excel upload configuration missing for '{model_config_key}'.")
        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    model_name = config['model']
    required_excel_headers = config.get('required_excel_headers', [])
    field_mapping = config.get('field_mapping', {})
    related_field_lookups = config.get('related_field_lookups', {})
    default_values = config.get('default_values', {})

    model_class = allif_main_models_registry.get(model_name)
    if not model_class:
        messages.error(request, f"Target model '{model_name}' not found in registry for Excel upload.")
        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    # Use an in-memory file to open with openpyxl
    file_content = io.BytesIO(excel_file.read())
    
    try:
        workbook = openpyxl.load_workbook(file_content)
        sheet = workbook.active
        
        # Read headers from the first row
        headers = [cell.value for cell in sheet[1]]
        
        # Validate required headers
        missing_headers = [h for h in required_excel_headers if h not in headers]
        if missing_headers:
            messages.error(request, f"Missing required Excel columns: {', '.join(missing_headers)}. Please ensure all mandatory columns are present.")
            return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                    kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

        successful_uploads = 0
        failed_rows = []

        # Start an atomic transaction
        with transaction.atomic():
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2): # Start from row 2 (after headers)
                row_data = dict(zip(headers, row))
                instance_data = {}
                row_errors = []

                # Apply default values first
                for default_field, default_value in default_values.items():
                    instance_data[default_field] = default_value

                for excel_header, cell_value in row_data.items():
                    # Map Excel header to model field name
                    model_field_name = field_mapping.get(excel_header, excel_header) # Use mapping or direct header

                    # Skip if model field doesn't exist or isn't in mapping and not a direct match
                    if not hasattr(model_class, model_field_name):
                        continue # Skip columns not relevant to the model

                    # Handle related fields (ForeignKeys)
                    if model_field_name in related_field_lookups:
                        lookup_info = related_field_lookups[model_field_name]
                        related_model_name = lookup_info['model']
                        lookup_field = lookup_info['lookup_field']
                        
                        related_model_class = allif_main_models_registry.get(related_model_name)
                        if not related_model_class:
                            row_errors.append(f"Configuration error: Related model '{related_model_name}' not found for field '{model_field_name}'.")
                            continue

                        try:
                            # Look up the related object
                            related_obj = related_model_class.objects.get(**{lookup_field: cell_value})
                            instance_data[model_field_name] = related_obj
                        except related_model_class.DoesNotExist:
                            row_errors.append(f"Related '{lookup_field}' '{cell_value}' not found for field '{excel_header}'.")
                        except Exception as e:
                            row_errors.append(f"Error looking up related '{excel_header}': {e}.")
                    else:
                        # Handle direct field values and type conversions
                        field_obj = model_class._meta.get_field(model_field_name)
                        
                        try:
                            # Attempt type conversion based on field type
                            if field_obj.get_internal_type() == 'DecimalField':
                                instance_data[model_field_name] = Decimal(str(cell_value or 0)) # Ensure string conversion before Decimal
                            elif field_obj.get_internal_type() in ['DateField', 'DateTimeField']:
                                if cell_value:
                                    if isinstance(cell_value, datetime.datetime): # openpyxl might read dates as datetime objects
                                        instance_data[model_field_name] = timezone.make_aware(cell_value) if field_obj.get_internal_type() == 'DateTimeField' else cell_value.date()
                                    else: # Assume string format
                                        # Try common date formats
                                        parsed_date = None
                                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%d/%m/%Y']:
                                            try:
                                                parsed_date = datetime.datetime.strptime(str(cell_value), fmt)
                                                break
                                            except ValueError:
                                                pass
                                        if parsed_date:
                                            instance_data[model_field_name] = timezone.make_aware(parsed_date) if field_obj.get_internal_type() == 'DateTimeField' else parsed_date.date()
                                        else:
                                            raise ValueError(f"Invalid date format for '{excel_header}': '{cell_value}'. Expected YYYY-MM-DD or similar.")
                                else:
                                    instance_data[model_field_name] = None
                            elif field_obj.get_internal_type() == 'BooleanField':
                                if isinstance(cell_value, str):
                                    instance_data[model_field_name] = cell_value.lower() in ['true', '1', 'yes', 'on']
                                else:
                                    instance_data[model_field_name] = bool(cell_value)
                            else:
                                instance_data[model_field_name] = cell_value # Default to direct assignment
                        except ValueError as ve:
                            row_errors.append(f'Data type error for "{excel_header}" ("{cell_value}"): {ve}.')
                        except Exception as e:
                            row_errors.append(f'Unexpected error processing "{excel_header}" ("{cell_value}"): {e}.')

                # Add multi-tenancy fields (company, division, branch, department)
                if hasattr(model_class, 'company'):
                    instance_data['company'] = CommonCompanyDetailsModel.all_objects.get(pk=company_id) # Assuming company ID is PK
                if hasattr(model_class, 'division') and allif_data.get("logged_in_user_has_divisional_access"):
                    instance_data['division'] = CommonDivisionsModel.all_objects.get(pk=allif_data.get("logged_user_division").id)
                if hasattr(model_class, 'branch') and allif_data.get("logged_in_user_has_branches_access"):
                    instance_data['branch'] = CommonBranchesModel.all_objects.get(pk=allif_data.get("logged_user_branch").id)
                if hasattr(model_class, 'department') and allif_data.get("logged_in_user_has_departmental_access"):
                    instance_data['department'] = CommonDepartmentsModel.all_objects.get(pk=allif_data.get("logged_user_department").id)
                
                # Try to create/update the instance
                if not row_errors:
                    try:
                        # For updates, you'd need a unique identifier from Excel (e.g., 'Code' or 'Part Number')
                        # For simplicity, this example always creates new instances.
                        # If you need update logic, you'd add a 'unique_lookup_field' to config
                        # and try to get_or_create or update_or_create.
                        instance = model_class(**instance_data)
                        instance.full_clean() # Run model validation
                        instance.save()
                        successful_uploads += 1
                    except ValidationError as e:
                        row_errors.append(f'Model validation error: {e.message_dict}')
                    except IntegrityError as e:
                        row_errors.append(f'Database integrity error (e.g., duplicate unique field): {e}.')
                    except Exception as e:
                        row_errors.append(f'Failed to save row: {e}.')
                
                if row_errors:
                    failed_rows.append({'row_number': row_idx, 'errors': row_errors, 'data': row_data})

        # Provide feedback
        if successful_uploads > 0:
            messages.success(request, f'Successfully uploaded {successful_uploads} records for {model_name}.')
        
        if failed_rows:
            error_msg = f'Failed to upload {len(failed_rows)} records for {model_name}. Details:'
            for fail in failed_rows:
                pass
                #error_msg += f'\nRow {fail['row_number']}: {'; '.join(fail['errors'])}. Data: {fail['data']}'
            messages.error(request, error_msg)
            # You might want to log these detailed errors to a file as well

        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

    except Exception as e:
        messages.error(request, f'An unexpected error occurred during Excel processingqqqqqqqqqqqqqqqqqq: {e}')
        import logging
        logger = logging.getLogger(__name__)
        logger.exception(f'CRITICAL ERROR: Excel upload failed for {model_config_key}: {e}')
        return redirect(reverse(f'allifmaalcommonapp:{success_redirect_url_name}', 
                                kwargs={'allifusr': user_slug, 'allifslug': company_slug}))

###############3 below is for function redirections... ####################
# --- NEW: Generic Sector-Based Redirect Helper Function ---
def allif_redirect_based_on_sector(request: HttpRequest, allif_data: dict, redirect_type: str) -> HttpResponse:
    """
    Redirects the user to the appropriate app's home or dashboard based on their
    company's sector.
    """
    company_entity = allif_data.get("main_sbscrbr_entity")
    user_slug = allif_data.get("usrslg")
    company_slug = allif_data.get("compslg")

    if not company_entity:
        # This case should ideally be caught earlier by common_shared_data or decorators
        # but as a fallback, redirect to company creation or decision point
        return redirect('allifmaalcommonapp:commonAddnewEntity', allifusr=user_slug)

    sector_name = str(company_entity.sector) # Convert sector object to string name

    target_urls = allif_sector_redirect_map.get(sector_name)

    if target_urls and redirect_type in target_urls:
        url_name = target_urls[redirect_type]
        return redirect(url_name, allifusr=user_slug, allifslug=company_slug)
    else:
        # If sector not found in map or redirect_type not defined for sector
        #logger.warning(f"No redirect URL found for sector '{sector_name}' and type '{redirect_type}'. Redirecting to CommonDecisionPoint.")
        messages.error(request, f"Your company's sector '{sector_name}' is not configured for this action. Please contact support.")
        return redirect('allifmaalcommonapp:CommonDecisionPoint')

# ... (rest of your utils.py code) ...



# --- Central helper function for Listing and Adding items ---
def allif_list_add_handler(request, model_class, form_class, template_name, title, redirect_view_name, *allifargs, **allifkwargs):
    """
    Handles a view that lists all objects and also contains a form for adding a new one.
    """
    allif_data = common_shared_data(request)
    allifqueryset = model_class.all_objects.all()
    form = form_class()
    if request.method == 'POST':
        form = form_class(request.POST or None)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = allif_data.get("logged_in_user")
            obj.save()
            return redirect(redirect_view_name,allifusr=allif_data.get("usrslg"),allifslug=allif_data.get("compslg"))
        else:
            error_message = form.errors
            allifcontext = {"error_message": error_message, "title": title}
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    context={"title": title,"form": form,"allifqueryset": allifqueryset,}
    return render(request, template_name, context)

def allif_edit_handler(request, model_class, form_class, pk, template_name, title, redirect_view_name, *allifargs, **allifkwargs):
    """
    Handles a view for editing an existing object.
    """
    allif_data = common_shared_data(request)
    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
    form = form_class(instance=allifquery)
    
    if request.method == 'POST':
        form = form_class(request.POST, instance=allifquery)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = allif_data.get("logged_in_user")
            obj.save()
            return redirect(redirect_view_name,allifusr=allif_data.get("usrslg"),allifslug=allif_data.get("compslg"))
        else:
            error_message = form.errors
            allifcontext = {"error_message": error_message, "title": title}
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    
    context={'form': form,"allifquery":allifquery,"title": title,}
    
    # Check if a queryset for the page is needed for display purposes
    if 'allifqueryset' in allifkwargs:
        context['allifqueryset'] = allifkwargs['allifqueryset']
    
    return render(request, template_name, context)

# --- Central helper function for Viewing item details ---
def allif_detail_handler(request, model_class, pk, template_name, title, *allifargs, **allifkwargs):
    """
    Fetches a single object and its related queryset for a detail view.
    """
    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
    context = {"allifquery": allifquery,"title": title,}

    if 'related_queryset' in allifkwargs:
        context['allifqueryset'] = allifkwargs['related_queryset']

    return render(request, template_name, context)

# --- Central helper function for Deleting items ---
def allif_deleting_hanlder(request, model_class, pk, redirect_view_name, *allifargs, **allifkwargs):
    """
    Finds and deletes an object, then redirects.
    """
    allif_data = common_shared_data(request)
    obj = get_object_or_404(model_class.all_objects, pk=pk)
    obj.delete()
    return redirect(redirect_view_name,allifusr=allif_data.get("usrslg"),allifslug=allif_data.get("compslg"))

# --- Central helper function for displaying a delete confirmation page ---
def allif_delete_confirm_handler(request, model_class, pk, template_name, title, *allifargs, **allifkwargs):
    """
    Finds an object and displays a confirmation page before deletion.
    """
    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
    context = {"title": title,"allifquery": allifquery,}
    return render(request, template_name, context)




######################### DIVISIONS, BRANCHES, DEPARTMENTS, OPERATION YEARS, OPERATION TERMS ##################
def allif_filtered_queryset(model_class, allif_data):
    """
    A helper to get a queryset for a given model, filtered based on the
    user's hierarchical access level (universal, divisional, etc.).
    """
    main_company = allif_data.get("main_sbscrbr_entity")

    if allif_data.get("logged_in_user_has_universal_access"):
        return model_class.all_objects.filter(company=main_company)
    elif allif_data.get("logged_in_user_has_divisional_access"):
        return model_class.all_objects.filter(
            company=main_company,
            division=allif_data.get("logged_user_division")
        )
    elif allif_data.get("logged_in_user_has_branches_access"):
        return model_class.all_objects.filter(
            company=main_company,
            division=allif_data.get("logged_user_division"),
            branch=allif_data.get("logged_user_branch")
        )
    elif allif_data.get("logged_in_user_has_departmental_access"):
        return model_class.all_objects.filter(
            company=main_company,
            division=allif_data.get("logged_user_division"),
            branch=allif_data.get("logged_user_branch"),
            department=allif_data.get("logged_user_department")
        )
    # If no specific access is found, return an empty queryset
    return model_class.all_objects.none()

# --- Central helper function for listing objects ---
def allif_list_view_handler(request, model_class, template_name, title):
    """
    Handles a view that lists all objects with access control.
    """
    allif_data = common_shared_data(request)
    allifqueryset = allif_filtered_queryset(model_class, allif_data)
    
    context = {
        "title": title,
        "allifqueryset": allifqueryset,
    }
    return render(request, template_name, context)

# --- Central helper function for Adding items with access control ---
def allif_add_view_handler(request, model_class, form_class, template_name, title, redirect_view_name, extra_context=None, specific_field_mapping=None):
    """
    Handles adding a new object, with access control and dynamic form initialization.
    'specific_field_mapping' can be used to set specific fields on the model instance before saving.
    """
    allif_data = common_shared_data(request)
    main_company = allif_data.get("main_sbscrbr_entity")
    
    # Check if form needs to be initialized with 'company' argument
    if hasattr(form_class, '__init__') and 'allifmaalparameter' in form_class.__init__.__code__.co_varnames:
        form = form_class(main_company)
    else:
        form = form_class()
    
    if request.method == 'POST':
        if hasattr(form_class, '__init__') and 'allifmaalparameter' in form_class.__init__.__code__.co_varnames:
            form = form_class(main_company, request.POST, request.FILES)
        else:
            form = form_class(request.POST, request.FILES)
        
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = allif_data.get("logged_in_user")
            obj.company = main_company

            # Set specific fields if provided
            if specific_field_mapping:
                for attr, val in specific_field_mapping.items():
                    setattr(obj, attr, val)

            obj.save()
            return redirect(redirect_view_name,
                            allifusr=allif_data.get("usrslg"),
                            allifslug=allif_data.get("compslg"))
        else:
            error_message = form.errors
            allifcontext = {"error_message": error_message, "title": title}
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    
    context = {
        "title": title,
        "form": form,
    }
    if extra_context:
        context.update(extra_context)

    return render(request, template_name, context)

# --- Central helper function for Editing items with access control ---
def allif_edit_view_handler(request, model_class, form_class, pk, template_name, title, redirect_view_name, extra_context=None):
    """
    Handles editing an existing object with access control.
    """
    allif_data = common_shared_data(request)
    main_company = allif_data.get("main_sbscrbr_entity")
    update_allifquery = get_object_or_404(model_class.all_objects, pk=pk)

    # Check if form needs to be initialized with 'company' argument
    if hasattr(form_class, '__init__') and 'allifmaalparameter' in form_class.__init__.__code__.co_varnames:
        form = form_class(main_company, instance=update_allifquery)
    else:
        form = form_class(instance=update_allifquery)

    if request.method == 'POST':
        if hasattr(form_class, '__init__') and 'allifmaalparameter' in form_class.__init__.__code__.co_varnames:
            form = form_class(main_company, request.POST, request.FILES, instance=update_allifquery)
        else:
            form = form_class(request.POST, request.FILES, instance=update_allifquery)
        
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = allif_data.get("logged_in_user")
            obj.save()
            return redirect(redirect_view_name,
                            allifusr=allif_data.get("usrslg"),
                            allifslug=allif_data.get("compslg"))
        else:
            error_message = form.errors
            allifcontext = {"error_message": error_message, "title": title}
            return render(request, 'allifmaalcommonapp/error/form-error.html', allifcontext)
    
    context = {
        'title': title,
        'form': form,
    }
    if extra_context:
        context.update(extra_context)
    
    return render(request, template_name, context)

# --- Central helper function for Viewing item details with related objects ---
def allif_detail_view_handler(request, model_class, pk, template_name, title, related_model=None):
    """
    Handles a detail view, fetching a single object and a related queryset based on access control.
    """
    allif_data = common_shared_data(request)
    main_company = allif_data.get("main_sbscrbr_entity")

    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
    relatedqueryset = None

    if related_model:
        # Dynamically filter related queryset based on user access
        related_queryset = allif_filtered_queryset(related_model, allif_data)
        
        # Filter the related queryset by the parent object
        if model_class is CommonDivisionsModel:
            related_queryset = related_queryset.filter(division=allifquery)
        elif model_class is CommonBranchesModel:
            related_queryset = related_queryset.filter(branch=allifquery)
        elif model_class is CommonDepartmentsModel:
            related_queryset = related_queryset.filter(department=allifquery)

    context = {
        "allifquery": allifquery,
        "title": title,
        "relatedqueryset": relatedqueryset,
    }
    return render(request, template_name, context)

# --- Central helper function for deletion logic ---
def allif_delete_view_handler(request, model_class, pk, redirect_view_name):
    """
    Handles the actual deletion of an object and redirects.
    """
    allif_data = common_shared_data(request)
    obj = get_object_or_404(model_class.all_objects, pk=pk)
    obj.delete()
    return redirect(redirect_view_name,
                    allifusr=allif_data.get("usrslg"),
                    allifslug=allif_data.get("compslg"))

# --- Central helper function for delete confirmation ---
def allif_delete_confirm_view_handler(request, model_class, pk, template_name, title):
    """
    Displays a confirmation page before deletion.
    """
    allifquery = get_object_or_404(model_class.all_objects, pk=pk)
    context = {
        "title": title,
        "allifquery": allifquery,
    }
    return render(request, template_name, context)


###############3 pdf reports generators for common reports like creditors, debtors, stock etc...########3
# This is the new centralized function that handles all the repetitive logic.
def allif_pdf_reports_generator(request, title, filename, template_path, allifqueryset, extra_context=None):
    allif_data = common_shared_data(request)
    scopes = CommonCompanyScopeModel.all_objects.filter(company=allif_data.get("main_sbscrbr_entity")).order_by('-date')[:4]

    context={"title": title,"scopes": scopes,"main_sbscrbr_entity": allif_data.get("main_sbscrbr_entity"),
        "allifqueryset": allifqueryset,}
    if extra_context:
        context.update(extra_context)

    # Prepare the HttpResponse for the PDF file
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{filename}"'

    # Render the HTML template
    template = get_template(template_path)
    html = template.render(context)

    # Create the PDF using xhtml2pdf
    try:
        pisa_status = pisa.CreatePDF(html, dest=response)
    except Exception as ex:
        # A more robust error message for the user
        return HttpResponse(f"Something went wrong while generating the PDF: {ex}")

    # If an error occurred during PDF creation, show a fallback HTML page
    if pisa_status.err:
        return HttpResponse(f'We had some errors during PDF creation. HTML: <pre>{html}</pre>')
        
    return response

# Refactored views using the new centralized function